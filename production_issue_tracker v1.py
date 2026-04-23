"""
AMAT Production Issue Tracker
Requires: pip install customtkinter openpyxl pillow
"""

import customtkinter as ctk
import tkinter as tk
import sys
from tkinter import ttk, messagebox, filedialog
import sqlite3, os, datetime, calendar
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from PIL import Image, ImageFont
from pilmoji import Pilmoji
import tksheet


# ══════════════════════════════════════════════════════════════════════════════
#  ✏️  EASY CUSTOMISATION — edit these blocks to change the look & feel
# ══════════════════════════════════════════════════════════════════════════════

# --- Colours -----------------------------------------------------------------
#  Key         What it controls
#  ─────────── ────────────────────────────────────────────────────────────────
#  bg          Window / page background
#  surface     Card / panel backgrounds (table, filter bar)
#  panel       Muted strip backgrounds (stats bar, section bars)
#  header      Top nav bar + section header bars
#  accent      Primary button fill, selected calendar day
#  accent_h    Hover shade for accent buttons
#  entry_bg    Text-entry & textbox backgrounds
#  border      Input borders, thin lines
#  text        Main body text
#  subtle      Muted / helper text, filter labels
#  open        "Open" status colour (red)
#  closed      "Closed" status colour (green)
#  stripe      Alternating row stripe in the table
#  sel         Selected row highlight / today-highlight in calendar
C = {
    "bg"      : "#e2ecf7",
    "surface" : "#f0f6fc",
    "panel"   : "#cddcee",
    "header"  : "#007ea5",
    "accent"  : "#2471c4",
    "accent_h": "#1a5ba0",
    "entry_bg": "#f5f9ff",
    "border"  : "#96b5d2",
    "text"    : "#0e1d2e",
    "subtle"  : "#557799",
    "open"    : "#b22222",
    "clarif"  : "#d87300",
    "closed"  : "#1a6e38",
    "stripe"  : "#e6f0fb",
    "sel"     : "#b5d2ee",
}

# --- Fonts -------------------------------------------------------------------
#  family    Font face used throughout the UI
#  size_sm   Small text  (filter labels, helper hints)
#  size_md   Normal text (inputs, buttons, table rows)
#  size_lg   Dialog sub-headings
#  size_xl   Page / dialog main headings
F = {
    "family" : "Montserrat",
    "size_sm": 10,
    "size_md": 11,
    "size_lg": 13,
    "size_xl": 15,
}

# --- Window ------------------------------------------------------------------
WIN_TITLE = "AMAT Production Issue Tracker"
WIN_SIZE  = "1420x820"   # initial width x height
WIN_MIN   = (1000, 620)  # minimum width, height

# --- Table columns -----------------------------------------------------------
#  Each tuple: (column_id, header_label, pixel_width, word_wrap?)
#  Reorder, resize, or hide columns by editing this list.
#  NOTE: Row numbers are handled by tksheet's built-in row index (left margin)
TABLE_COLS = [
    ("date",     "Date",             125, False),
    ("system",   "System No.",       230, True ),
    ("family",   "Product Family",   130, False),
    ("type",     "Issue Type",       165, True),
    ("sps",      "SPS No.",           145, False),
    ("ncr",      "NCR No.",          145, False),
    ("status",   "Status",            145, False),
    ("desc",     "Issue Description",530, True ),
    ("solution", "Solution",         530, True ),
    ("sol_date", "Sol. Date",         145, False),
    ("crf",      "CRF",               100, False),
    ("esw",      "ESW",               100, False),
    ("scv",      "SCV",               100, False),
    ("remarks",  "Remarks",          530, True ),
]

# --- Dropdown options --------------------------------------------------------
FAMILIES    = ["FEP", "DDP", "ETCH"]
ISSUE_TYPES = ["BOM Error", "Document Discrepancy", "Document Error",
               "Missing Document", "Design Error", "Others"]

# --- Paths -------------------------------------------------------------------
if getattr(sys, 'frozen', False):
    # If running as an .exe
    _HERE = os.path.dirname(sys.executable)
    _BUNDLE = sys._MEIPASS # Internal path for bundled files like logo.png
else:
    # If running as a normal .py script
    _HERE = os.path.dirname(os.path.abspath(__file__))
    _BUNDLE = _HERE

DB_PATH   = os.path.join(_HERE, "production_issues.db")
LOGO_PATH = os.path.join(_HERE, "logo.png")

# --- Derived filter lists (don't edit) ---------------------------------------
MONTH_NAMES = ["All"] + [datetime.date(2000, m, 1).strftime("%B") for m in range(1, 13)]
YEAR_OPTS   = ["All"] + [str(y) for y in range(2022, datetime.date.today().year + 3)]


# ══════════════════════════════════════════════════════════════════════════════
#  DATABASE
# ══════════════════════════════════════════════════════════════════════════════
def _db():
    return sqlite3.connect(DB_PATH)

def _row_factory(cursor, row):
    return dict(zip([d[0] for d in cursor.description], row))

def init_db():
    with _db() as c:
        c.execute("""
            CREATE TABLE IF NOT EXISTS issues (
                id             INTEGER PRIMARY KEY AUTOINCREMENT,
                date_reported  TEXT DEFAULT '',
                system_number  TEXT DEFAULT '',
                product_family TEXT DEFAULT '',
                issue_type     TEXT DEFAULT '',
                issue_desc     TEXT DEFAULT '',
                sps_number     TEXT DEFAULT '',
                ncr_number     TEXT DEFAULT '',
                status         TEXT DEFAULT 'Open',
                solution       TEXT DEFAULT '',
                solution_date  TEXT DEFAULT '',
                crf            TEXT DEFAULT '',
                esw            TEXT DEFAULT '',
                scv            TEXT DEFAULT '',
                remarks        TEXT DEFAULT '',
                created_at     TEXT DEFAULT (datetime('now','localtime'))
            )
        """)
        existing = {r[1] for r in c.execute("PRAGMA table_info(issues)")}
        for col, defn in [
            ("ncr_number",    "TEXT DEFAULT ''"),
            ("solution",      "TEXT DEFAULT ''"),
            ("solution_date", "TEXT DEFAULT ''"),
            ("crf",           "TEXT DEFAULT ''"),
            ("esw",           "TEXT DEFAULT ''"),
            ("scv",           "TEXT DEFAULT ''"),
            ("remarks",       "TEXT DEFAULT ''"),
        ]:
            if col not in existing:
                c.execute(f"ALTER TABLE issues ADD COLUMN {col} {defn}")

def fetch_issues(status_f="All", family_f="All", itype_f="All",
                 month_f="All", year_f="All"):
    q    = """SELECT id, date_reported, system_number, product_family, issue_type,
                     issue_desc, sps_number, ncr_number, status,
                     solution, solution_date, crf, esw, scv, remarks, created_at
              FROM issues WHERE 1=1"""
    args = []
    if status_f != "All": q += " AND status=?";         args.append(status_f)
    if family_f != "All": q += " AND product_family=?"; args.append(family_f)
    if itype_f  != "All": q += " AND issue_type=?";     args.append(itype_f)
    if month_f  != "All":
        q += " AND CAST(strftime('%m', date_reported) AS INTEGER)=?"
        args.append(datetime.datetime.strptime(month_f, "%B").month)
    if year_f   != "All":
        q += " AND strftime('%Y', date_reported)=?";    args.append(year_f)
    q += " ORDER BY date_reported DESC, id DESC"
    with _db() as c:
        c.row_factory = _row_factory
        return c.execute(q, args).fetchall()

def fetch_by_id(db_id):
    with _db() as c:
        c.row_factory = _row_factory
        return c.execute("SELECT * FROM issues WHERE id=?", (db_id,)).fetchone()

def insert_issue(date_reported, system_number, product_family, issue_type,
                 issue_desc, sps_number, ncr_number):
    with _db() as c:
        cur = c.execute(
            """INSERT INTO issues
               (date_reported, system_number, product_family, issue_type,
                issue_desc, sps_number, ncr_number)
               VALUES (?,?,?,?,?,?,?)""",
            (date_reported, system_number, product_family, issue_type,
             issue_desc, sps_number, ncr_number))
        return cur.lastrowid

def update_issue(db_id, date_reported, system_number, product_family, issue_type,
                 issue_desc, sps_number, ncr_number, status,
                 solution, solution_date, crf, esw, scv, remarks):
    with _db() as c:
        c.execute("""
            UPDATE issues SET
                date_reported=?, system_number=?, product_family=?, issue_type=?,
                issue_desc=?, sps_number=?, ncr_number=?, status=?,
                solution=?, solution_date=?, crf=?, esw=?, scv=?, remarks=?
            WHERE id=?""",
            (date_reported, system_number, product_family, issue_type,
             issue_desc, sps_number, ncr_number, status,
             solution, solution_date, crf, esw, scv, remarks, db_id))

def delete_by_ids(ids):
    with _db() as c:
        c.executemany("DELETE FROM issues WHERE id=?", [(i,) for i in ids])

def get_counts():
    with _db() as c:
        total  = c.execute("SELECT COUNT(*) FROM issues").fetchone()[0]
        open_  = c.execute("SELECT COUNT(*) FROM issues WHERE status='Open'").fetchone()[0]
        closed = c.execute("SELECT COUNT(*) FROM issues WHERE status='Closed'").fetchone()[0]
        clarif = c.execute("SELECT COUNT(*) FROM issues WHERE status='Clarification'").fetchone()[0]
    return total, open_, closed, clarif


# ══════════════════════════════════════════════════════════════════════════════
#  EXCEL EXPORT
# ══════════════════════════════════════════════════════════════════════════════
def _fmt_date(d):
    try:    return datetime.datetime.strptime(d, "%Y-%m-%d").strftime("%d %b %Y")
    except: return d or ""

def _xl_border():
    s = Side(style="thin", color="B0C8E0")
    return Border(left=s, right=s, top=s, bottom=s)

def _auto_row_height(text, col_width_chars, base_pt=13, min_pt=18, padding=4):
    if not text:
        return min_pt
    lines = sum(max(1, -(-len(p) // max(1, col_width_chars)))
                for p in str(text).splitlines())
    return max(min_pt, lines * base_pt + padding)

def export_excel(rows, filepath, customer_mode=False):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Production Issues"

    if customer_mode:
        headers    = ["Date Reported", "System Number", "Product Family",
                      "Issue Type", "Issue Description", "SPS Number", "Remarks"]
        col_widths = [14, 22, 14, 18, 60, 13, 35]
        left_cols  = {4, 6}
    else:
        headers    = ["Date Reported", "System Number", "Product Family", "Issue Type",
                      "Issue Description", "SPS Number", "NCR Number", "Status",
                      "Solution", "Solution Date", "CRF", "ESW", "SCV", "Remarks"]
        col_widths = [14, 22, 14, 18, 55, 13, 13, 10, 45, 14, 12, 12, 12, 35]
        left_cols  = {4, 8, 13}

    hdr_fill = PatternFill("solid", fgColor="1B5DA8")
    hdr_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    ctr = Alignment(horizontal="center", vertical="top", wrap_text=True)
    lft = Alignment(horizontal="left",   vertical="top", wrap_text=True)

    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = hdr_font; cell.fill = hdr_fill
        cell.alignment = ctr; cell.border = _xl_border()
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 26

    open_fill   = PatternFill("solid", fgColor="FDEDEC")
    closed_fill = PatternFill("solid", fgColor="EAFAF1")
    alt_fill    = PatternFill("solid", fgColor="EBF4FF")

    for ri, row in enumerate(rows, 2):
        fill = (open_fill   if row["status"] == "Open"   else
                closed_fill if row["status"] == "Closed" else alt_fill)
        values = (
            [_fmt_date(row["date_reported"]), row["system_number"],
             row["product_family"], row["issue_type"], row["issue_desc"],
             row["sps_number"], row["remarks"]]
            if customer_mode else
            [_fmt_date(row["date_reported"]), row["system_number"],
             row["product_family"], row["issue_type"], row["issue_desc"],
             row["sps_number"], row["ncr_number"], row["status"],
             row["solution"], _fmt_date(row["solution_date"]),
             row["crf"], row["esw"], row["scv"], row["remarks"]]
        )
        row_h = max(
            (_auto_row_height(val, cw)
             for ci0, (val, cw) in enumerate(zip(values, col_widths))
             if ci0 in left_cols),
            default=18
        )
        ws.row_dimensions[ri].height = row_h
        for ci, (val, cw) in enumerate(zip(values, col_widths), 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.fill = fill; cell.border = _xl_border()
            cell.font = Font(name="Calibri", size=10)
            cell.alignment = lft if (ci - 1) in left_cols else ctr

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    wb.save(filepath)


# ══════════════════════════════════════════════════════════════════════════════
#  UI PRIMITIVES
# ══════════════════════════════════════════════════════════════════════════════
def _font(**kw):
    return ctk.CTkFont(family=F["family"], **kw)

def _lbl(parent, text, bold=False, size=None, color=None, **kw):
    return ctk.CTkLabel(parent, text=text, anchor="w",
                        text_color=color or C["text"],
                        font=_font(size=size or F["size_md"],
                                   weight="bold" if bold else "normal"), **kw)

def _entry(parent, ph="", width=None, state="normal"):
    kw = dict(fg_color=C["entry_bg"], text_color=C["text"], border_color=C["border"],
              corner_radius=8, placeholder_text=ph, state=state,
              font=_font(size=F["size_md"]))
    if width:
        kw["width"] = width
    e = ctk.CTkEntry(parent, **kw)
    stack = []
    def _push(*_):
        v = e.get()
        if not stack or stack[-1] != v:
            stack.append(v)
            if len(stack) > 200: stack.pop(0)
    def _undo(_):
        if len(stack) > 1:
            stack.pop(); e.delete(0, "end"); e.insert(0, stack[-1])
        return "break"
    e.bind("<KeyRelease>", _push)
    e.bind("<Control-z>",  _undo)
    return e

def _textbox(parent, h=80):
    tb = ctk.CTkTextbox(parent, height=h, fg_color=C["entry_bg"],
                        text_color=C["text"], border_color=C["border"],
                        border_width=1, corner_radius=8,
                        font=_font(size=F["size_md"]),
                        wrap="word")
    tb._textbox.configure(undo=True, maxundo=-1)
    return tb

def _btn(parent, text, cmd, outline=False, width=120, color=None, radius=10, **kw):
    if outline:
        return ctk.CTkButton(parent, text=text, command=cmd, width=width,
                             corner_radius=radius, fg_color="transparent",
                             border_width=1, border_color=C["accent"],
                             text_color=C["accent"], hover_color=C["panel"],
                             font=_font(size=F["size_md"]), **kw)
    return ctk.CTkButton(parent, text=text, command=cmd, width=width,
                         corner_radius=radius, fg_color=color or C["accent"],
                         hover_color=C["accent_h"], text_color="white",
                         font=_font(size=F["size_md"]), **kw)

def _section_bar(parent, title, row):
    f = ctk.CTkFrame(parent, fg_color=C["header"], corner_radius=0, height=24)
    f.grid(row=row, column=0, columnspan=2, sticky="ew", pady=(10, 2))
    _lbl(f, f"  {title}", size=F["size_sm"], bold=True, color="white"
         ).pack(side="left", padx=8, pady=2)

def _optmenu(parent, vals, var, width=None):
    return ToggleDropdown(parent, vals, var, width=width)

def _popup_xy(widget):
    widget.update_idletasks()
    return widget.winfo_rootx(), widget.winfo_rooty() + widget.winfo_height() + 4


# ══════════════════════════════════════════════════════════════════════════════
#  TOGGLE DROPDOWN
# ══════════════════════════════════════════════════════════════════════════════
class ToggleDropdown(ctk.CTkFrame):
    def __init__(self, master, values, variable, width=None):
        super().__init__(master, fg_color="transparent")
        self._values = values
        self._var    = variable
        self._popup  = None
        w = width or 150
        self._btn = ctk.CTkButton(
            self, text=self._label(), width=w, anchor="w",
            fg_color=C["accent"], hover_color=C["accent_h"],
            text_color="white", corner_radius=8,
            font=_font(size=F["size_md"]),
            command=self._toggle)
        self._btn.pack(fill="both", expand=True)
        variable.trace_add("write", lambda *_: self._btn.configure(text=self._label()))

    def _label(self):
        return f"  {self._var.get()}  ▾"

    def _toggle(self):
        if self._popup and self._popup.winfo_exists():
            self._close()
        else:
            self._show()

    def _show(self):
        x, y = _popup_xy(self._btn)
        bw   = self._btn.winfo_width()

        # Cap popup height so it never runs off-screen; add scroll when needed
        row_h     = 32
        max_rows  = 10                          # show at most 10 rows before scrolling
        n         = len(self._values)
        visible   = min(n, max_rows)
        pop_h     = visible * row_h

        self._popup = tk.Toplevel(self)
        self._popup.overrideredirect(True)
        self._popup.geometry(f"{bw}x{pop_h}+{x}+{y}")
        self._popup.configure(bg=C["border"])
        self._popup.lift()

        # Scrollable canvas so all items are reachable even when the list is long
        canvas = tk.Canvas(self._popup, bg=C["surface"], highlightthickness=0,
                           width=bw-2, height=pop_h)
        vsb    = tk.Scrollbar(self._popup, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=vsb.set)

        if n > max_rows:
            vsb.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)

        inner = tk.Frame(canvas, bg=C["surface"])
        win_id = canvas.create_window((0, 0), window=inner, anchor="nw")

        def _on_inner_resize(e):
            canvas.configure(scrollregion=canvas.bbox("all"))
            canvas.itemconfig(win_id, width=canvas.winfo_width())
        inner.bind("<Configure>", _on_inner_resize)
        canvas.bind("<Configure>",
                    lambda e: canvas.itemconfig(win_id, width=e.width))

        # Mouse-wheel scrolling inside the popup
        def _wheel(e):
            canvas.yview_scroll(int(-1 * (e.delta / 120)), "units")
        canvas.bind("<MouseWheel>", _wheel)
        inner.bind("<MouseWheel>",  _wheel)

        for val in self._values:
            rf = tk.Frame(inner, bg=C["surface"], cursor="hand2", height=row_h)
            rf.pack(fill="x"); rf.pack_propagate(False)
            lb = tk.Label(rf, text=f"  {val}", bg=C["surface"], fg=C["text"],
                          font=(F["family"], F["size_sm"]), anchor="w")
            lb.pack(fill="both", expand=True)
            for w2 in (rf, lb):
                w2.bind("<Enter>",    lambda e, r=rf, l=lb:
                        (r.configure(bg=C["sel"]),     l.configure(bg=C["sel"])))
                w2.bind("<Leave>",    lambda e, r=rf, l=lb:
                        (r.configure(bg=C["surface"]), l.configure(bg=C["surface"])))
                w2.bind("<Button-1>", lambda e, v=val: self._pick(v))
                w2.bind("<MouseWheel>", _wheel)

        self._popup.after(50, lambda: self._popup.focus_force()
                          if self._popup and self._popup.winfo_exists() else None)
        self._popup.bind("<FocusOut>", self._on_focusout)

    def _on_focusout(self, _):
        if not (self._popup and self._popup.winfo_exists()): return
        bx, by = self._btn.winfo_rootx(), self._btn.winfo_rooty()
        bw, bh = self._btn.winfo_width(),  self._btn.winfo_height()
        mx, my = self._popup.winfo_pointerxy()
        self._close(skip_toggle=(bx <= mx <= bx+bw and by <= my <= by+bh))

    def _pick(self, v):
        self._var.set(v); self._close()

    def _close(self, skip_toggle=False):
        if self._popup and self._popup.winfo_exists():
            self._popup.destroy()
        self._popup = None
        if skip_toggle:
            self._btn.configure(state="disabled")
            self.after(150, lambda: self._btn.configure(state="normal"))


# ══════════════════════════════════════════════════════════════════════════════
#  CALENDAR WIDGET
#
#  Built on a plain tk.Toplevel with grab_set() so that:
#    • All mouse/keyboard events are captured by the popup window.
#    • A click anywhere outside is detected via the event's root coordinates —
#      no FocusOut / focus_get() trickery needed.
#    • Child buttons receiving focus never accidentally dismiss the popup.
# ══════════════════════════════════════════════════════════════════════════════
class CalendarPopup(tk.Toplevel):
    def __init__(self, parent, anchor_widget, initial=None):
        super().__init__(parent)
        self.overrideredirect(True)
        self.configure(bg=C["surface"])
        self._result = None
        self._view   = initial or datetime.date.today()
        self._sel    = initial

        self._build()
        self._render()

        # Position below the anchor button once layout is committed
        self.update_idletasks()
        x, y = _popup_xy(anchor_widget)
        self.geometry(f"+{x}+{y}")
        self.lift()

        # Grab all events — clicks outside will still generate Button-1 on self
        self.grab_set()
        self.bind("<Button-1>", self._on_click)
        self.bind("<Escape>",   lambda _: self._close())

    # ── Layout ────────────────────────────────────────────────────────────────
    def _build(self):
        nav = tk.Frame(self, bg=C["header"])
        nav.pack(fill="x")
        for side, txt, cmd in [("left", "◀", self._prev), ("right", "▶", self._next)]:
            tk.Button(nav, text=txt, bg=C["header"], fg="white", relief="flat",
                      bd=0, padx=8, font=(F["family"], 11, "bold"),
                      activebackground=C["accent_h"], activeforeground="white",
                      command=cmd).pack(side=side, pady=5)
        self._hdr = tk.Label(nav, text="", bg=C["header"], fg="white",
                             font=(F["family"], 11, "bold"), width=18)
        self._hdr.pack(side="left", expand=True)

        dow = tk.Frame(self, bg=C["panel"])
        dow.pack(fill="x")
        for i, d in enumerate(["Mo", "Tu", "We", "Th", "Fr", "Sa", "Su"]):
            tk.Label(dow, text=d, width=4, bg=C["panel"], fg=C["subtle"],
                     font=(F["family"], 9, "bold")).grid(row=0, column=i, padx=3, pady=3)

        self._grid = tk.Frame(self, bg=C["surface"])
        self._grid.pack(padx=6, pady=4)

        foot = tk.Frame(self, bg=C["surface"])
        foot.pack(fill="x", pady=(0, 6))
        tk.Button(foot, text="Today", bg=C["panel"], fg=C["text"], relief="flat",
                  font=(F["family"], 9), activebackground=C["border"],
                  command=self._pick_today).pack()

    def _render(self):
        for w in self._grid.winfo_children():
            w.destroy()
        y, m  = self._view.year, self._view.month
        today = datetime.date.today()
        self._hdr.config(text=f"{calendar.month_name[m]}  {y}")
        for r, week in enumerate(calendar.monthcalendar(y, m)):
            for cc, day in enumerate(week):
                if day == 0:
                    tk.Label(self._grid, text="", width=4, bg=C["surface"]
                             ).grid(row=r, column=cc, padx=2, pady=2)
                    continue
                d   = datetime.date(y, m, day)
                sel = d == self._sel
                tod = d == today
                bg  = C["accent"] if sel else (C["sel"] if tod else C["surface"])
                fg  = "white" if sel else C["text"]
                fw  = "bold"  if sel or tod else "normal"
                tk.Button(self._grid, text=str(day), width=4, bg=bg, fg=fg,
                          relief="flat", bd=0, font=(F["family"], 10, fw),
                          activebackground=C["accent"], activeforeground="white",
                          command=lambda dd=d: self._pick(dd)
                          ).grid(row=r, column=cc, padx=2, pady=2)

    # ── Interaction ───────────────────────────────────────────────────────────
    def _on_click(self, event):
        """Close if the click landed outside this popup's bounding box."""
        wx, wy = self.winfo_rootx(), self.winfo_rooty()
        ww, wh = self.winfo_width(),  self.winfo_height()
        if not (wx <= event.x_root <= wx + ww and wy <= event.y_root <= wy + wh):
            self._close()

    def _pick(self, d):
        self._result = d.strftime("%Y-%m-%d")
        self._close()

    def _pick_today(self):
        self._result = datetime.date.today().strftime("%Y-%m-%d")
        self._close()

    def _close(self):
        self.grab_release()
        self.destroy()

    def _prev(self):
        y, m = self._view.year, self._view.month - 1
        if m == 0: m, y = 12, y - 1
        self._view = datetime.date(y, m, 1); self._render()

    def _next(self):
        y, m = self._view.year, self._view.month + 1
        if m == 13: m, y = 1, y + 1
        self._view = datetime.date(y, m, 1); self._render()

    def result(self):
        return self._result


# ══════════════════════════════════════════════════════════════════════════════
#  DATE FIELD  (read-only entry + calendar button)
# ══════════════════════════════════════════════════════════════════════════════
class DateField(ctk.CTkFrame):
    def __init__(self, master, initial=None, nullable=False, **kw):
        super().__init__(master, fg_color="transparent", **kw)
        self._val      = initial or ("" if nullable else datetime.date.today().strftime("%Y-%m-%d"))
        self._var      = tk.StringVar(value=_fmt_date(self._val))
        self._nullable = nullable

        ctk.CTkEntry(self, textvariable=self._var, state="readonly", width=116,
                     fg_color=C["entry_bg"], text_color=C["text"],
                     border_color=C["border"], corner_radius=8).pack(side="left")
        self._cal_btn = ctk.CTkButton(self, text="📅", width=32, corner_radius=8,
                                      fg_color=C["accent"], hover_color=C["accent_h"],
                                      command=self._open)
        self._cal_btn.pack(side="left", padx=(4, 0))
        if nullable:
            ctk.CTkButton(self, text="✕", width=26, corner_radius=8,
                          fg_color=C["panel"], hover_color=C["border"],
                          text_color=C["text"],
                          command=self._clear).pack(side="left", padx=(2, 0))

    def _open(self):
        try:
            init = datetime.datetime.strptime(self._val, "%Y-%m-%d").date() if self._val else None
        except ValueError:
            init = None
        popup = CalendarPopup(self, anchor_widget=self._cal_btn, initial=init)
        self.wait_window(popup)
        if popup.result():
            self._val = popup.result()
            self._var.set(_fmt_date(self._val))

    def _clear(self):
        self._val = ""; self._var.set("")

    def get(self):    return self._val
    def set(self, v): self._val = v; self._var.set(v)


# ══════════════════════════════════════════════════════════════════════════════
#  COPY-PASTE TEXT GENERATOR
# ══════════════════════════════════════════════════════════════════════════════
class CopyPasteDialog(ctk.CTkToplevel):
    def __init__(self, parent, system_number, issue_desc, sps_number):
        super().__init__(parent)
        self.title("Copy-Paste Text")
        self.geometry("640x620")
        self.resizable(True, True)
        self.grab_set()
        self.configure(fg_color=C["bg"])
        self._build(system_number, issue_desc, sps_number)

    def _build(self, sys_num, desc, sps_num):
        systems_block = sys_num.strip()
        email_text = (
            f"Hi KB / Zach,\n\n"
            f"SPS {sps_num} submitted for the following issue(s):\n\n"
            f"{systems_block}\n\n"
            f"{desc}\n\n"
            f"Best regards,"
        )
        desc_of_def = f"{systems_block}\n\n{desc}"
        det_dispos  = f"AMAT SPS {sps_num} submitted."

        outer = ctk.CTkScrollableFrame(self, fg_color=C["bg"], corner_radius=0)
        outer.pack(fill="both", expand=True)

        def _block(title, content, h=120):
            _lbl(outer, title, bold=True, size=F["size_sm"], color=C["header"]
                 ).pack(anchor="w", padx=18, pady=(14, 2))
            tb = _textbox(outer, h=h)
            tb.pack(fill="x", padx=18)
            tb.insert("1.0", content)
            def _copy():
                self.clipboard_clear()
                self.clipboard_append(tb.get("1.0", "end").rstrip())
                btn.configure(text="Copied ✓")
                self.after(1500, lambda: btn.configure(text="Copy"))
            btn = _btn(outer, "Copy", _copy, outline=True, width=80, radius=7)
            btn.pack(anchor="e", padx=18, pady=(3, 0))

        _block("📧  Email body",                            email_text,  h=170)
        _block("📋  NCR — Desc. of Def. / Req. for Change", desc_of_def, h=140)
        _block("📋  NCR — Det. Dispos. / Reas. for Change", det_dispos,  h=60)
        _btn(self, "Close", self.destroy, outline=True, width=100).pack(pady=14)


# ══════════════════════════════════════════════════════════════════════════════
#  NEW ISSUE DIALOG
# ══════════════════════════════════════════════════════════════════════════════
class NewIssueDialog(ctk.CTkToplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title("Log New Issue")
        self.geometry("640x720")
        self.resizable(False, True)
        self.grab_set()
        self.configure(fg_color=C["bg"])
        self._build()

    def _build(self):
        self.grid_columnconfigure(1, weight=1)
        P = dict(padx=22, pady=4)
        r = 0

        _lbl(self, "New Issue", bold=True, size=F["size_xl"]).grid(
            row=r, column=0, columnspan=2, padx=22, pady=(18, 10), sticky="w"); r += 1

        _lbl(self, "Date Reported *").grid(row=r, column=0, **P, sticky="w")
        self.dt = DateField(self)
        self.dt.grid(row=r, column=1, **P, sticky="w"); r += 1

        _lbl(self, "System Number(s) *").grid(row=r, column=0, **P, sticky="nw")
        sf = ctk.CTkFrame(self, fg_color="transparent")
        sf.grid(row=r, column=1, **P, sticky="ew")
        self.sys = _textbox(sf, h=60); self.sys.pack(fill="x")
        _lbl(sf, "One system per line for multiple systems.",
             size=F["size_sm"]-1, color=C["subtle"]).pack(anchor="w", pady=(1, 0)); r += 1

        _lbl(self, "Product Family *").grid(row=r, column=0, **P, sticky="w")
        self.fam = ctk.StringVar(value=FAMILIES[0])
        _optmenu(self, FAMILIES, self.fam).grid(row=r, column=1, **P, sticky="ew"); r += 1

        _lbl(self, "Issue Type *").grid(row=r, column=0, **P, sticky="w")
        self.ityp = ctk.StringVar(value=ISSUE_TYPES[0])
        _optmenu(self, ISSUE_TYPES, self.ityp).grid(row=r, column=1, **P, sticky="ew"); r += 1

        _lbl(self, "Issue Description *").grid(row=r, column=0, **P, sticky="nw")
        df = ctk.CTkFrame(self, fg_color="transparent")
        df.grid(row=r, column=1, **P, sticky="ew")
        self.desc = _textbox(df, h=130); self.desc.pack(fill="x")
        _lbl(df, "Separate multiple issues with a blank line.",
             size=F["size_sm"]-1, color=C["subtle"]).pack(anchor="w", pady=(1, 0)); r += 1

        _lbl(self, "SPS Number *").grid(row=r, column=0, **P, sticky="w")
        self.sps = _entry(self, "e.g. 752766")
        self.sps.grid(row=r, column=1, **P, sticky="ew"); r += 1

        _lbl(self, "NCR Number").grid(row=r, column=0, **P, sticky="w")
        self.ncr = _entry(self, "e.g. NCR281174  (fill after creating in Agile)")
        self.ncr.grid(row=r, column=1, **P, sticky="ew"); r += 1

        bf = ctk.CTkFrame(self, fg_color="transparent")
        bf.grid(row=r, column=0, columnspan=2, pady=18)
        _btn(bf, "Cancel",                   self.destroy,    outline=True, width=90
             ).pack(side="left", padx=6)
        _btn(bf, "Generate Copy-Paste Text", self._generate,
             color=C["subtle"], width=210).pack(side="left", padx=6)
        _btn(bf, "Submit",                   self._submit,    width=120
             ).pack(side="left", padx=6)

    def _generate(self):
        CopyPasteDialog(self,
                        self.sys.get("1.0", "end").strip(),
                        self.desc.get("1.0", "end").strip(),
                        self.sps.get().strip())

    def _submit(self):
        d    = self.dt.get().strip()
        s    = self.sys.get("1.0", "end").strip()
        desc = self.desc.get("1.0", "end").strip()
        sps  = self.sps.get().strip()
        if not all([d, s, desc, sps]):
            messagebox.showerror("Missing Fields",
                                 "Date Reported, System Number, Issue Description "
                                 "and SPS Number are required.", parent=self)
            return
        insert_issue(d, s, self.fam.get(), self.ityp.get(), desc,
                     sps, self.ncr.get().strip())
        self.destroy()


# ══════════════════════════════════════════════════════════════════════════════
#  MANAGE / EDIT DIALOG
# ══════════════════════════════════════════════════════════════════════════════
class ManageDialog(ctk.CTkToplevel):
    def __init__(self, parent, db_id):
        super().__init__(parent)
        self.db_id = db_id
        self.title("Edit Issue")
        self.geometry("720x880")
        self.resizable(True, True)
        self.grab_set()
        self.configure(fg_color=C["bg"])
        self.R = fetch_by_id(db_id)
        self._build()

    def _build(self):
        sf = ctk.CTkScrollableFrame(self, fg_color=C["bg"], corner_radius=0,
                                    scrollbar_button_color=C["accent"],
                                    scrollbar_button_hover_color=C["accent_h"])
        sf.pack(fill="both", expand=True)
        sf.grid_columnconfigure(1, weight=1)
        P = dict(padx=22, pady=4)
        R = self.R; r = 0

        title_sys = R["system_number"].splitlines()[0] if R["system_number"] else ""
        _lbl(sf, f"Issue — {title_sys}", bold=True, size=F["size_lg"]).grid(
            row=r, column=0, columnspan=2, padx=22, pady=(16, 2), sticky="w"); r += 1
        _lbl(sf, f"Created: {R['created_at']}", size=F["size_sm"], color=C["subtle"]).grid(
            row=r, column=0, columnspan=2, padx=22, sticky="w"); r += 1

        _section_bar(sf, "Basic Information", r); r += 1

        _lbl(sf, "Date Reported").grid(row=r, column=0, **P, sticky="w")
        self.dt = DateField(sf, initial=R["date_reported"])
        self.dt.grid(row=r, column=1, **P, sticky="w"); r += 1

        _lbl(sf, "System Number(s)").grid(row=r, column=0, **P, sticky="nw")
        sys_f = ctk.CTkFrame(sf, fg_color="transparent")
        sys_f.grid(row=r, column=1, **P, sticky="ew")
        self.sys = _textbox(sys_f, h=60); self.sys.pack(fill="x")
        self.sys.insert("1.0", R["system_number"])
        _lbl(sys_f, "One system per line for multiple systems.",
             size=F["size_sm"]-1, color=C["subtle"]).pack(anchor="w", pady=(1, 0)); r += 1

        _lbl(sf, "Product Family").grid(row=r, column=0, **P, sticky="w")
        self.fam = ctk.StringVar(value=R["product_family"])
        _optmenu(sf, FAMILIES, self.fam).grid(row=r, column=1, **P, sticky="ew"); r += 1

        _lbl(sf, "Issue Type").grid(row=r, column=0, **P, sticky="w")
        self.ityp = ctk.StringVar(value=R["issue_type"])
        _optmenu(sf, ISSUE_TYPES, self.ityp).grid(row=r, column=1, **P, sticky="ew"); r += 1

        _lbl(sf, "Issue Description").grid(row=r, column=0, **P, sticky="nw")
        df = ctk.CTkFrame(sf, fg_color="transparent")
        df.grid(row=r, column=1, **P, sticky="ew")
        self.desc = _textbox(df, h=100); self.desc.pack(fill="x")
        self.desc.insert("1.0", R["issue_desc"])
        _lbl(df, "Separate multiple issues with a blank line.",
             size=F["size_sm"]-1, color=C["subtle"]).pack(anchor="w", pady=(1, 0)); r += 1

        _lbl(sf, "SPS Number").grid(row=r, column=0, **P, sticky="w")
        self.sps = _entry(sf); self.sps.insert(0, R["sps_number"])
        self.sps.grid(row=r, column=1, **P, sticky="ew"); r += 1

        _lbl(sf, "NCR Number").grid(row=r, column=0, **P, sticky="w")
        self.ncr = _entry(sf); self.ncr.insert(0, R["ncr_number"])
        self.ncr.grid(row=r, column=1, **P, sticky="ew"); r += 1

        _lbl(sf, "Status").grid(row=r, column=0, **P, sticky="w")
        self.status = ctk.StringVar(value=R["status"])
        rf = ctk.CTkFrame(sf, fg_color="transparent")
        rf.grid(row=r, column=1, **P, sticky="w")
        for v, col in [("Open", C["open"]), ("Clarification", C["clarif"]), ("Closed", C["closed"])]:
            ctk.CTkRadioButton(rf, text=v, value=v, variable=self.status,
                               fg_color=col, hover_color=col, text_color=C["text"],
                               corner_radius=50, font=_font(size=F["size_md"])
                               ).pack(side="left", padx=10)
        r += 1

        _section_bar(sf, "Resolution", r); r += 1

        _lbl(sf, "Solution").grid(row=r, column=0, **P, sticky="nw")
        self.sol = _textbox(sf, h=80)
        self.sol.insert("1.0", R["solution"] or "")
        self.sol.grid(row=r, column=1, **P, sticky="ew"); r += 1

        _lbl(sf, "Solution Date").grid(row=r, column=0, **P, sticky="w")
        self.sol_dt = DateField(sf, initial=R["solution_date"] or None, nullable=True)
        self.sol_dt.grid(row=r, column=1, **P, sticky="w"); r += 1

        _section_bar(sf, "Change Controls", r); r += 1
        for attr, label, val in [("crf", "CRF", R["crf"]),
                                  ("esw", "ESW", R["esw"]),
                                  ("scv", "SCV", R["scv"])]:
            _lbl(sf, label).grid(row=r, column=0, **P, sticky="w")
            e = _entry(sf); e.insert(0, val or "")
            e.grid(row=r, column=1, **P, sticky="ew")
            setattr(self, attr, e); r += 1

        _section_bar(sf, "Remarks", r); r += 1
        _lbl(sf, "Remarks").grid(row=r, column=0, **P, sticky="nw")
        self.rmk = _textbox(sf, h=70)
        self.rmk.insert("1.0", R["remarks"] or "")
        self.rmk.grid(row=r, column=1, **P, sticky="ew"); r += 1

        bb = ctk.CTkFrame(self, fg_color=C["panel"], corner_radius=0, height=52)
        bb.pack(fill="x", side="bottom"); bb.pack_propagate(False)
        _btn(bb, "Delete",       self._delete, color="#a02020", width=100, radius=8
             ).pack(side="left",  padx=14, pady=10)
        _btn(bb, "Cancel",       self.destroy, outline=True,    width=90,  radius=8
             ).pack(side="right", padx=8,  pady=10)
        _btn(bb, "Save Changes", self._save,                    width=150, radius=8
             ).pack(side="right", padx=8,  pady=10)

    def _save(self):
        update_issue(self.db_id,
            self.dt.get(),
            self.sys.get("1.0", "end").strip(),
            self.fam.get(), self.ityp.get(),
            self.desc.get("1.0", "end").strip(),
            self.sps.get().strip(), self.ncr.get().strip(),
            self.status.get(),
            self.sol.get("1.0", "end").strip(),
            self.sol_dt.get(),
            self.crf.get().strip(), self.esw.get().strip(), self.scv.get().strip(),
            self.rmk.get("1.0", "end").strip())
        messagebox.showinfo("Saved", "Issue updated.", parent=self)
        self.destroy()

    def _delete(self):
        if messagebox.askyesno("Delete", "Permanently delete this issue?", parent=self):
            delete_by_ids([self.db_id]); self.destroy()


# ══════════════════════════════════════════════════════════════════════════════
#  MAIN WINDOW
# ══════════════════════════════════════════════════════════════════════════════
class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        ctk.set_appearance_mode("light")
        ctk.set_default_color_theme("blue")
        self.configure(fg_color=C["bg"])
        self.title(WIN_TITLE)
        self.geometry(WIN_SIZE)
        self.minsize(*WIN_MIN)
        self._id_map   = {}
        self._raw_text = {}
        self._build()
        self._refresh()

    def _build(self):
        self._build_header()
        self._build_stats()
        self._build_filters()
        self._build_toolbar()
        self._build_table()

    # ── Header ────────────────────────────────────────────────────────────────
    def _build_header(self):
        bar = tk.Frame(self, bg=C["header"], height=54)
        bar.pack(fill="x"); bar.pack_propagate(False)

        if os.path.exists(LOGO_PATH):
            try:
                from PIL import ImageTk
                img = Image.open(LOGO_PATH)
                img.thumbnail((96, 72), Image.LANCZOS)
                self._logo = ImageTk.PhotoImage(img)
                tk.Label(bar, image=self._logo, bg=C["header"],
                         borderwidth=0).pack(side="left", padx=(14, 6), pady=9)
            except Exception:
                pass

        tk.Label(bar, text=WIN_TITLE, bg=C["header"], fg="white",
                 font=(F["family"], F["size_xl"], "bold")
                 ).pack(side="left", padx=(0, 20), pady=(20,0))

        for txt, cmd, bold in [
            ("Refresh",      self._refresh,   False),
            ("Export Excel", self._export,    False),
            ("+ New Issue",  self._new_issue, True),
        ]:
            tk.Button(bar, text=txt,
                      bg="#174e8a" if bold else C["header"],
                      fg="white", relief="flat", bd=0, padx=14, pady=5,
                      activebackground=C["accent_h"], activeforeground="white",
                      font=(F["family"], F["size_md"], "bold" if bold else "normal"),
                      command=cmd).pack(side="right", padx=(0, 8), pady=12)

    # ── Stats bar ─────────────────────────────────────────────────────────────
    def _build_stats(self):
        bar = tk.Frame(self, bg=C["panel"], height=34)
        bar.pack(fill="x"); bar.pack_propagate(False)
        self._v_total  = tk.StringVar(value="Total: 0")
        self._v_open   = tk.StringVar(value="Open: 0")
        self._v_clarif = tk.StringVar(value="Clarification: 0")
        self._v_closed = tk.StringVar(value="Closed: 0")
        for var, col in [(self._v_total,  C["text"]),
                         (self._v_open,   C["open"]),
                         (self._v_clarif, C["clarif"]),
                         (self._v_closed, C["closed"])]:
            tk.Label(bar, textvariable=var, bg=C["panel"], fg=col,
                     font=(F["family"], F["size_md"], "bold")
                     ).pack(side="left", padx=18, pady=6)
            
        tk.Label(bar, text="Made by Sankar  |  v1.0",
                 bg=C["panel"], fg=C["subtle"],
                 font=(F["family"], F["size_sm"] - 1)).pack(side="right", padx=14)

    # ── Filters ───────────────────────────────────────────────────────────────
    def _build_filters(self):
        bar = ctk.CTkFrame(self, fg_color=C["surface"], corner_radius=0)
        bar.pack(fill="x", ipady=6)
        row = ctk.CTkFrame(bar, fg_color="transparent")
        row.pack(pady=(12,0), padx=14)
        _lbl(row, "Filters:", bold=True, size=F["size_sm"], color=C["subtle"]
             ).grid(row=0, column=0, padx=(0, 10))

        def flt(label, var, opts, col, w=None):
            _lbl(row, label, size=F["size_sm"], color=C["subtle"]
                 ).grid(row=0, column=col, padx=(0, 3))
            _optmenu(row, opts, var, width=w).grid(row=0, column=col+1, padx=(0, 12))
            var.trace_add("write", lambda *_: self._refresh())

        self._fs = ctk.StringVar(value="All")
        self._ff = ctk.StringVar(value="All")
        self._fi = ctk.StringVar(value="All")
        self._fm = ctk.StringVar(value="All")
        self._fy = ctk.StringVar(value="All")
        flt("Status", self._fs, ["All", "Open", "Clarification", "Closed"],  1, 100)
        flt("Family", self._ff, ["All"] + FAMILIES,         3, 100)
        flt("Issue",  self._fi, ["All"] + ISSUE_TYPES,      5, 175)
        flt("Month",  self._fm, MONTH_NAMES,                7, 120)
        flt("Year",   self._fy, YEAR_OPTS,                  9,  88)
        _btn(row, "Reset", self._reset_filters, outline=True, width=68, radius=8
             ).grid(row=0, column=11, padx=(4, 0))

    # ── Toolbar ───────────────────────────────────────────────────────────────
    def _build_toolbar(self):
        bar = ctk.CTkFrame(self, fg_color="transparent")
        bar.pack(fill="x", padx=14, pady=(10, 0))

        _btn(bar, "Delete Selected", self._mass_delete,
             color="#a02020", width=148, radius=8).pack(side="left")
        self._sel_lbl = _lbl(bar, "", size=F["size_sm"], color=C["subtle"])
        self._sel_lbl.pack(side="left", padx=10)
        
        # Status label for copy feedback
        self._status_lbl = _lbl(bar, "", size=F["size_sm"], color=C["accent"])
        self._status_lbl.pack(side="left", padx=10)

        # ── Search bar (right side of toolbar) ────────────────────────────────
        _lbl(bar, "Hold Ctrl / Shift to select multiple rows",
             size=F["size_sm"], color=C["subtle"]).pack(side="right", padx=(0, 16))

        # Match counter
        self._search_match_lbl = _lbl(bar, "", size=F["size_sm"], color=C["subtle"])
        self._search_match_lbl.pack(side="right", padx=(0, 6))

        # Clear button (×)
        self._search_clear_btn = _btn(bar, "✕", self._search_clear,
                                       outline=True, width=28, radius=6)
        self._search_clear_btn.pack(side="right", padx=(0, 2))
        self._search_clear_btn.pack_forget()   # hidden until there's a query

        # Search entry
        self._search_var = tk.StringVar()
        self._search_entry = ctk.CTkEntry(
            bar, textvariable=self._search_var,
            placeholder_text="🔍  Search all columns…",
            width=240, corner_radius=8,
            fg_color=C["entry_bg"], text_color=C["text"],
            border_color=C["border"],
            font=_font(size=F["size_md"]))
        self._search_entry.pack(side="right", padx=(0, 4))
        self._search_var.trace_add("write", lambda *_: self._search_apply())

        # Ctrl+F focuses the search bar from anywhere in the window
        self.bind_all("<Control-f>", lambda e: (
            self._search_entry.focus_set(),
            self._search_entry.select_range(0, "end")))
        # Escape clears and dismisses search
        self._search_entry.bind("<Escape>", lambda e: self._search_clear())

    # ── Table ─────────────────────────────────────────────────────────────────
    def _build_table(self):
        outer = ctk.CTkFrame(self, fg_color=C["surface"], corner_radius=10)
        outer.pack(fill="both", expand=True, padx=14, pady=8)

        # Extract column headers and widths
        col_hdrs = [c[1] for c in TABLE_COLS]
        col_wids = [c[2] for c in TABLE_COLS]
        self._col_hdrs = {c[0]: c[1] for c in TABLE_COLS}  # id → header text
        self._col_ids = [c[0] for c in TABLE_COLS]
        self._wrap_cols = {c[0] for c in TABLE_COLS if c[3]}
        
        # Create tksheet.Sheet widget with professional styling
        self.sheet = tksheet.Sheet(
            outer,
            data=[],  # Start with empty data (headers handled separately)
            column_width=150,  # Default column width
            height=400,
            width=800,
            theme="light",
            empty_horizontal=0,  # Remove extra horizontal padding
            empty_vertical=0,    # Remove extra vertical padding
            header_align="center",  # Center-align headers by default
            align="center"  # Center-align cells by default
        )
        
        # Set proper column names (not numbers)
        self.sheet.headers(col_hdrs)
        
        # Enable comprehensive bindings for professional spreadsheet feel
        self.sheet.enable_bindings((
            "copy",
            "select",
            "drag_select",
            "single_select",
            "multi_select",
            "row_select",
            "column_select",
            "ctrl_click_select",
            "shift_click_select",
            "column_width_resize",
            "row_height_resize",
            "double_click_column_resize"
        ))
        
        # Bind double-click to open issue
        self.sheet.extra_bindings([("double_click_table", self._on_double_click)])
        
        # Configure row index (left margin)
        self.sheet.row_index_width = 50
        self.sheet.show_row_index = True
        
        # Set column widths from TABLE_COLS
        for col_idx, (col_id, header, width, _) in enumerate(TABLE_COLS):
            self.sheet.column_width(col_idx, width=width)
        
        # Configure alignment for text-heavy columns (constructor already sets center-align default)
        # Set left alignment for text-heavy columns using column index list
        left_align_indices = []
        for col_idx, (col_id, _, _, _) in enumerate(TABLE_COLS):
            if col_id in ("desc", "solution", "remarks"):
                left_align_indices.append(col_idx)
        
        # Apply left alignment to these specific columns
        for col_idx in left_align_indices:
            self.sheet.align_columns(columns=col_idx, align='w')
        
        # Configure fonts and sizes
        self.sheet.font = (F["family"], F["size_md"])
        self.sheet.header_font = (F["family"], 11, "bold")
        
        # Apply comprehensive professional styling via set_options
        self.sheet.set_options(
            header_height=26,
            header_bg=C["header"],
            header_fg="white",
            table_bg=C["surface"],
            table_fg=C["text"],
            index_bg=C["panel"],
            index_fg=C["text"],
            grid_color=C["border"]
        )
        
        # Bind click events for copy feedback
        self.sheet.bind("<Button-1>", self._on_sheet_click, add="+")
        self.sheet.bind("<Button-3>", self._on_sheet_right_click, add="+")
        
        self.sheet.pack(fill="both", expand=True, padx=6, pady=6)
        
        # Store state for tracking selected rows
        self._selected_rows = set()
        self._sort_col = None
        self._sort_reverse = False
        self._flashing_cells = {}
    
    def _on_double_click(self, event=None):
        """Handle double-click to open issue for editing."""
        try:
            if event:
                # Extract row index from event tuple
                row = event[0] if isinstance(event, (list, tuple)) and len(event) > 0 else None
                if row and row > 0 and row in self._id_map:
                    db_id = self._id_map[row]
                    dlg = ManageDialog(self, db_id)
                    self.wait_window(dlg)
                    self._refresh()
                    return
        except Exception:
            pass
    
    def _on_sheet_click(self, event=None):
        """Handle sheet clicks: copy to clipboard and flash effect (respects Ctrl/Shift for multi-select)."""
        if not event or self.sheet.identify_region(event) != "body":
            return
        
        # Get the clicked cell/row
        row = self.sheet.identify_row(event, return_val="data")
        col = self.sheet.identify_column(event, return_val="data")
        
        if row is None or col is None or row == 0:  # Skip header row
            return
        
        # Check if Ctrl or Shift is pressed (for multi-select support)
        has_modifier = (event.state & 0x0004) or (event.state & 0x0001)  # Ctrl or Shift
        
        # Only perform copy/flash if this is a single click (no modifiers)
        if not has_modifier:
            try:
                cell_value = self.sheet.get_cell_value(row, col)
                if cell_value:
                    # Copy to clipboard
                    self.clipboard_clear()
                    self.clipboard_append(str(cell_value))
                    self._show_copy_feedback(str(cell_value))
                
                # Flash the cell
                self._flash_cell_tksheet(row, col)
            except Exception:
                pass
    
    def _on_sheet_right_click(self, event=None):
        """Handle right-click for context menu (optional)."""
        pass
    
    def _flash_cell_tksheet(self, row, col):
        """Display a brief flash (250ms) over the clicked cell."""
        try:
            # Highlight cell with accent color using cells parameter
            self.sheet.highlight_cells(cells=[(row, col)], bg=C["accent"], fg="white", overwrite=False)
            
            # Store for later clearing
            self._flashing_cells[(row, col)] = True
            
            # Clear highlight after 250ms
            def clear_highlight():
                try:
                    self.sheet.dehighlight_cells(cells=[(row, col)])
                    if (row, col) in self._flashing_cells:
                        del self._flashing_cells[(row, col)]
                except Exception:
                    pass
            
            self.after(250, clear_highlight)
        except Exception:
            pass

    def _show_copy_feedback(self, text):
        """Display a brief 'Copied' message in the status label and auto-hide it."""
        # Truncate very long text for display
        display_text = text[:50] + "..." if len(text) > 50 else text
        self._status_lbl.configure(text=f"✓ Copied: {display_text}")
        
        # Auto-clear the status after 2 seconds
        if hasattr(self, "_status_after_id"):
            self.after_cancel(self._status_after_id)
        self._status_after_id = self.after(2000, lambda: self._status_lbl.configure(text=""))

    # ── Search ────────────────────────────────────────────────────────────────
    def _search_apply(self, *_):
        """Filter rows by search query and refresh table."""
        query = self._search_var.get().strip().lower()

        # Show/hide the clear button
        if query:
            self._search_clear_btn.pack(side="right", padx=(0, 2),
                                         before=self._search_entry)
        else:
            self._search_clear_btn.pack_forget()

        # Refresh to apply filters
        self._refresh()
        
        # Update match counter
        n = len(self._raw_text)
        self._search_match_lbl.configure(
            text=f"{n} match{'es' if n != 1 else ''}" if n else "No matches",
            text_color=C["subtle"] if n else C["open"])

    def _search_clear(self):
        """Clear the search query and restore all rows."""
        self._search_var.set("")
        self._search_entry.focus_set()

    # ── Data ──────────────────────────────────────────────────────────────────
    def _refresh(self, *_):
        self._id_map.clear()
        self._raw_text.clear()
        self._selected_rows.clear()
        
        # Clear search and sort state on a full refresh
        self._search_var.set("")
        self._search_match_lbl.configure(text="")
        # Clear sort state
        self._sort_col = None
        self._sort_reverse = False

        col_ids = self._col_ids
        rows = fetch_issues(self._fs.get(), self._ff.get(),
                           self._fi.get(), self._fm.get(), self._fy.get())
        
        # Build 2D array for tksheet: data rows only (headers handled by self.sheet.headers())
        table_data = []
        
        for row_num, row in enumerate(rows, start=1):
            status = row["status"]
            
            raw = {
                "date"    : _fmt_date(row["date_reported"]),
                "system"  : row["system_number"]  or "",
                "family"  : row["product_family"],
                "type"    : row["issue_type"],
                "sps"     : row["sps_number"],
                "ncr"     : row["ncr_number"],
                "status"  : status,
                "desc"    : row["issue_desc"]      or "",
                "solution": row["solution"]         or "",
                "sol_date": _fmt_date(row["solution_date"]),
                "crf"     : row["crf"],
                "esw"     : row["esw"],
                "scv"     : row["scv"],
                "remarks" : row["remarks"]          or "",
            }
            
            # Build row data in column order (# column removed from TABLE_COLS)
            row_data = [raw.get(c, "") for c in col_ids]
            table_data.append(row_data)
            
            # Map tksheet row index (1-based) to database ID
            self._id_map[row_num] = row["id"]
            self._raw_text[row_num] = raw
        
        # Update sheet with new data (no header row in data)
        self.sheet.set_sheet_data(table_data, reset_col_positions=False)
        
        # Re-apply column widths from TABLE_COLS
        for col_idx, (col_id, header, width, _) in enumerate(TABLE_COLS):
            self.sheet.column_width(col_idx, width=width)
        
        # Set left alignment for text-heavy columns after data is loaded (constructor set center-align default)
        left_align_indices = []
        for col_idx, (col_id, _, _, _) in enumerate(TABLE_COLS):
            if col_id in ("desc", "solution", "remarks"):
                left_align_indices.append(col_idx)
        
        for col_idx in left_align_indices:
            self.sheet.align_columns(columns=col_idx, align='w')
        
        # Update statistics
        t, o, cl, cf = get_counts()
        self._v_total.set(f"Total: {t}")
        self._v_open.set(f"Open: {o}")
        self._v_clarif.set(f"Clarification: {cf}")
        self._v_closed.set(f"Closed: {cl}")
        self._sel_lbl.configure(text="")
        
        # Auto-size rows for wrapping text (CRUCIAL: at very end of refresh)
        try:
            self.sheet.set_all_row_heights_to_auto(redraw=True)
        except Exception:
            # Fallback if auto-height fails
            pass

    def _reset_filters(self):
        for v in [self._fs, self._ff, self._fi, self._fm, self._fy]:
            v.set("All")

    def _sort_by_column(self, col_idx):
        """Sort by column (preserves row numbers #)."""
        if self._sort_col == col_idx:
            self._sort_reverse = not self._sort_reverse
        else:
            self._sort_col = col_idx
            self._sort_reverse = False
        
        # Re-fetch and re-sort the data
        self._refresh()

    def _on_select(self, _):
        """Handle row selection (handled by click binding)."""
        pass

    def _deselect_if_outside(self, event):
        """Deselect if click is outside the sheet."""
        pass

    # ── Actions ───────────────────────────────────────────────────────────────
    def _new_issue(self):
        dlg = NewIssueDialog(self)
        self.wait_window(dlg); self._refresh()

    def _open_issue(self, event=None):
        """Open the first selected row's issue details or from double-click."""
        try:
            # If called from double-click event, extract row index
            if event and hasattr(event, "data"):
                row = event.data[0] if isinstance(event.data, (list, tuple)) else event.data
                if row and row > 0 and row in self._id_map:
                    db_id = self._id_map[row]
                    dlg = ManageDialog(self, db_id)
                    self.wait_window(dlg)
                    self._refresh()
                    return
        except Exception:
            pass
        
        # Fallback: open first selected row or first row in table
        if self._selected_rows:
            row_idx = min(self._selected_rows)
            if row_idx in self._id_map:
                db_id = self._id_map[row_idx]
                dlg = ManageDialog(self, db_id)
                self.wait_window(dlg); self._refresh()
        elif self._id_map:
            # If no explicit selection, open the first row
            first_row = min(self._id_map.keys())
            db_id = self._id_map[first_row]
            dlg = ManageDialog(self, db_id)
            self.wait_window(dlg); self._refresh()

    def _mass_delete(self):
        """Delete all selected rows."""
        if not self._selected_rows:
            messagebox.showwarning("No Selection",
                                   "Please select at least one row using Ctrl+click or Shift+click.",
                                   parent=self)
            return
        
        if not messagebox.askyesno("Confirm Delete",
                                   f"Delete {len(self._selected_rows)} issue(s)? This cannot be undone.",
                                   parent=self):
            return
        
        # Get the database IDs for selected rows
        ids_to_delete = [self._id_map[row_idx] for row_idx in self._selected_rows 
                         if row_idx in self._id_map]
        
        if ids_to_delete:
            delete_by_ids(ids_to_delete)
            messagebox.showinfo("Deleted", f"Deleted {len(ids_to_delete)} issue(s).", parent=self)
            self._refresh()

    def _export(self):
        rows = fetch_issues(self._fs.get(), self._ff.get(),
                            self._fi.get(), self._fm.get(), self._fy.get())
        if not rows:
            messagebox.showinfo("No Data", "No issues match the current filters.",
                                parent=self); return
        customer_mode = False
        if self._fs.get() == "Open":
            answer = messagebox.askyesnocancel(
                "Export Type",
                "You are exporting Open issues.\n\n"
                "Export as Customer Report?\n"
                "(Yes = customer view, no internal fields)\n"
                "(No  = full internal log)",
                parent=self)
            if answer is None: return
            customer_mode = answer
        suffix = "_Customer" if customer_mode else ""
        path = filedialog.asksaveasfilename(
            parent=self, defaultextension=".xlsx",
            filetypes=[("Excel Workbook", "*.xlsx")],
            initialfile=f"AMAT_Issues{suffix}_{datetime.date.today()}.xlsx")
        if not path: return
        try:
            export_excel(rows, path, customer_mode=customer_mode)
            if messagebox.askyesno("Exported",
                                   f"Saved {len(rows)} rows.\n\nOpen file now?",
                                   parent=self):
                os.startfile(path)
        except Exception as e:
            messagebox.showerror("Export Error", str(e), parent=self)


# ══════════════════════════════════════════════════════════════════════════════
#  ENTRY POINT
# ══════════════════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    init_db()
    App().mainloop()
