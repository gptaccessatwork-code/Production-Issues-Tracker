"""
AMAT Production Issue Tracker v3.0
Requires: pip install PySide6 openpyxl pillow
"""

from __future__ import annotations

import calendar
import datetime
import os
import sqlite3
import sys
from collections import defaultdict

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from PySide6.QtCore import QDate, QSignalBlocker, QSize, QPoint, QRectF, Qt, QTimer, Signal
from PySide6.QtGui import QGuiApplication
from PySide6.QtGui import QColor, QFont, QFontMetrics, QIcon, QKeySequence, QPainter, QPen, QPixmap, QTextCharFormat
from PySide6.QtWidgets import (
    QAbstractItemView,
    QApplication,
    QCalendarWidget,
    QComboBox,
    QDialog,
    QFileDialog,
    QFrame,
    QGridLayout,
    QHeaderView,
    QHBoxLayout,
    QLabel,
    QLineEdit,
    QMainWindow,
    QMessageBox,
    QMenu,
    QPushButton,
    QPlainTextEdit,
    QRadioButton,
    QScrollArea,
    QTabWidget,
    QTableWidget,
    QTableWidgetItem,
    QStyledItemDelegate,
    QSizePolicy,
    QToolButton,
    QVBoxLayout,
    QWidget,
)


# -----------------------------------------------------------------------------
#  EASY CUSTOMIZATION
# -----------------------------------------------------------------------------

C = {
    "bg": "#0f172a",
    "surface": "#111c33",
    "surface_2": "#16213b",
    "panel": "#1b2947",
    "header": "#0b1220",
    "accent": "#f59e0b",
    "accent_h": "#d97706",
    "entry_bg": "#0d1629",
    "border": "#2a3a5d",
    "text": "#e5eefc",
    "subtle": "#8ca0c4",
    "open": "#fb7185",
    "clarif": "#f59e0b",
    "closed": "#34d399",
    "stripe": "#12203a",
    "sel": "#21365f",
}

F = {
    "family": "Segoe UI",
    "size_sm": 10,
    "size_md": 11,
    "size_lg": 13,
    "size_xl": 15,
}

WIN_TITLE = "AMAT Production Issue Tracker"
WIN_SIZE = (1420, 820)
WIN_MIN = (1000, 620)

TABLE_COLS = [
    ("date", "Date", 125, False),
    ("system", "System No.", 230, True),
    ("family", "Product Family", 130, False),
    ("type", "Issue Type", 165, True),
    ("sps", "SPS No.", 145, False),
    ("ncr", "NCR No.", 145, False),
    ("status", "Status", 145, False),
    ("desc", "Issue Description", 530, True),
    ("solution", "Solution", 530, True),
    ("sol_date", "Sol. Date", 145, False),
    ("crf", "CRF", 100, False),
    ("esw", "ESW", 100, False),
    ("scv", "SCV", 100, False),
    ("remarks", "Remarks", 530, True),
]

FAMILIES = ["FEP", "DDP", "ETCH", "MDP", "EPI", "SPOTBUY/SPARES"]
ISSUE_TYPES = [
    "BOM Error",
    "Document Discrepancy",
    "Document Error",
    "Missing Document",
    "Design Error",
    "Request for Deviation",
    "Others",
]
TRACKER_TYPES = ["Engineering", "Material"]

if getattr(sys, "frozen", False):
    _HERE = os.path.dirname(sys.executable)
else:
    _HERE = os.path.dirname(os.path.abspath(__file__))

# Use the shared drive database so the app can run locally while reading/writing
# the central issue tracker data.
DB_PATH = r"S:\Engineering\003 AMAT Products\AMAT Production Issues\production_issues.db"
LOGO_PATH = os.path.join(_HERE, "logo.png")

MONTH_NAMES = ["All"] + [datetime.date(2000, m, 1).strftime("%B") for m in range(1, 13)]
YEAR_OPTS = ["All"] + [str(y) for y in range(2022, datetime.date.today().year + 3)]
TREND_MONTHS = [calendar.month_abbr[i] for i in range(1, 13)]
BULK_IMPORT_COLORS = ["#60a5fa", "#34d399", "#f59e0b", "#f472b6", "#a78bfa", "#22d3ee", "#f87171", "#cbd5e1"]
TREND_ISSUE_COLORS = {
    "BOM Error": "#34d399",
    "Document Discrepancy": "#60a5fa",
    "Document Error": "#f59e0b",
    "Missing Document": "#f472b6",
    "Design Error": "#a78bfa",
    "Request for Deviation": "#22d3ee",
    "Others": "#fb7185",
}
BULK_IMPORT_HEADERS = [
    "Date Reported",
    "System Number",
    "Product Family",
    "Issue Type",
    "Issue Description",
    "SPS Number",
    "NCR Number",
    "Status",
    "Solution",
    "Solution Date",
    "CRF",
    "ESW",
    "SCV",
    "Remarks",
    "Tracker Type",
]
BULK_IMPORT_TEMPLATE_PREFIX = "PIT_Bulk_Import_Template"


# -----------------------------------------------------------------------------
#  DATABASE
# -----------------------------------------------------------------------------

def _db():
    return sqlite3.connect(DB_PATH)


def _row_factory(cursor, row):
    return dict(zip([d[0] for d in cursor.description], row))


def _normalize_key(value):
    return "".join(ch for ch in str(value or "").strip().lower() if ch.isalnum())


def _clean_text(value):
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return str(value).strip()


def _coerce_date_to_iso(value):
    if not value:
        return ""
    if isinstance(value, datetime.datetime):
        return value.date().strftime("%Y-%m-%d")
    if isinstance(value, datetime.date):
        return value.strftime("%Y-%m-%d")
    text = _clean_text(value)
    if not text:
        return ""
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y", "%d %b %Y", "%d %B %Y"):
        try:
            return datetime.datetime.strptime(text, fmt).strftime("%Y-%m-%d")
        except Exception:
            pass
    return text


def _resolve_choice(value, allowed_values):
    text = _clean_text(value)
    if not text:
        return ""
    norm = _normalize_key(text)
    for allowed in allowed_values:
        if _normalize_key(allowed) == norm:
            return allowed
    return text


def _available_issue_years(tracker_type=None):
    q = "SELECT DISTINCT strftime('%Y', date_reported) AS y FROM issues WHERE date_reported<>''"
    args = []
    if tracker_type in TRACKER_TYPES:
        q += " AND tracker_type=?"
        args.append(tracker_type)
    q += " ORDER BY y DESC"
    with _db() as c:
        years = [r[0] for r in c.execute(q, args).fetchall() if r and r[0]]
    if not years:
        years = [str(datetime.date.today().year)]
    return years


def _bulk_import_template_path():
    return os.path.join(_HERE, f"{BULK_IMPORT_TEMPLATE_PREFIX}.xlsx")


def create_bulk_import_template(filepath):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bulk Import"
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(BULK_IMPORT_HEADERS))}1"

    header_fill = PatternFill("solid", fgColor="1B5DA8")
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, header in enumerate(BULK_IMPORT_HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        ws.column_dimensions[get_column_letter(col_idx)].width = {
            "Date Reported": 14,
            "System Number": 24,
            "Product Family": 16,
            "Issue Type": 20,
            "Issue Description": 55,
            "SPS Number": 14,
            "NCR Number": 14,
            "Status": 14,
            "Solution": 55,
            "Solution Date": 14,
            "CRF": 12,
            "ESW": 12,
            "SCV": 12,
            "Remarks": 35,
            "Tracker Type": 16,
        }.get(header, 18)

    ws.row_dimensions[1].height = 24

    dv_family = DataValidation(type="list", formula1=f'"{",".join(FAMILIES)}"', allow_blank=True)
    dv_issue = DataValidation(type="list", formula1=f'"{",".join(ISSUE_TYPES)}"', allow_blank=True)
    dv_status = DataValidation(type="list", formula1='"Open,Clarification,Closed"', allow_blank=True)
    dv_tracker = DataValidation(type="list", formula1=f'"{",".join(TRACKER_TYPES)}"', allow_blank=True)
    ws.add_data_validation(dv_family)
    ws.add_data_validation(dv_issue)
    ws.add_data_validation(dv_status)
    ws.add_data_validation(dv_tracker)
    dv_family.add("C2:C5000")
    dv_issue.add("D2:D5000")
    dv_status.add("H2:H5000")
    dv_tracker.add("O2:O5000")

    thin_wrap = Alignment(horizontal="left", vertical="top", wrap_text=True)
    date_fmt = "yyyy-mm-dd"
    for row_idx in range(2, 5001):
        ws.cell(row_idx, 1).number_format = date_fmt
        ws.cell(row_idx, 5).alignment = thin_wrap
        ws.cell(row_idx, 9).alignment = thin_wrap
        ws.cell(row_idx, 10).number_format = date_fmt
        ws.cell(row_idx, 14).alignment = thin_wrap

    center_cols = [2, 3, 4, 6, 7, 8, 11, 12, 13, 15]
    for row_idx in range(2, 5001):
        for col_idx in center_cols:
            ws.cell(row_idx, col_idx).alignment = Alignment(horizontal="center", vertical="center")

    tips = wb.create_sheet("Instructions")
    tips.column_dimensions["A"].width = 92
    for idx, text in enumerate(
        [
            "Fill one issue per row on the Bulk Import sheet.",
            "Leave any field blank if you do not want to populate it.",
            "Status defaults to Open when left blank.",
            "Tracker Type defaults to the active workspace when left blank.",
            "Date Reported and Solution Date accept Excel dates or ISO dates like 2026-07-10.",
            "Issue Description, Solution, and Remarks wrap text automatically.",
        ],
        1,
    ):
        tips.cell(idx, 1, text)

    wb.save(filepath)


def import_bulk_issues(filepath, default_tracker_type="Engineering"):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = None
    header_map = {}
    required_keys = [_normalize_key(h) for h in [
        "Date Reported",
        "System Number",
        "Product Family",
        "Issue Type",
        "Issue Description",
        "SPS Number",
    ]]
    for candidate in wb.worksheets:
        for row in candidate.iter_rows(min_row=1, max_row=min(10, candidate.max_row), values_only=True):
            if not row:
                continue
            candidate_map = {_normalize_key(v): idx for idx, v in enumerate(row) if _normalize_key(v)}
            if candidate_map and all(key in candidate_map for key in required_keys):
                ws = candidate
                header_map = candidate_map
                break
        if ws is not None:
            break

    if ws is None or not header_map:
        raise ValueError("Could not find a header row in the workbook.")

    imported = 0
    skipped = []
    inserts = []
    for excel_row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or all(cell is None or _clean_text(cell) == "" for cell in row):
            continue

        def cell(name):
            idx = header_map.get(_normalize_key(name))
            if idx is None or idx >= len(row):
                return None
            return row[idx]

        d = _coerce_date_to_iso(cell("Date Reported"))
        sys_num = _clean_text(cell("System Number"))
        family = _resolve_choice(cell("Product Family"), FAMILIES)
        issue_type = _resolve_choice(cell("Issue Type"), ISSUE_TYPES)
        issue_desc = _clean_text(cell("Issue Description"))
        sps = _clean_text(cell("SPS Number"))
        ncr = _clean_text(cell("NCR Number"))
        status = _resolve_choice(cell("Status"), ["Open", "Clarification", "Closed"]) or "Open"
        solution = _clean_text(cell("Solution"))
        solution_date = _coerce_date_to_iso(cell("Solution Date"))
        crf = _clean_text(cell("CRF"))
        esw = _clean_text(cell("ESW"))
        scv = _clean_text(cell("SCV"))
        remarks = _clean_text(cell("Remarks"))
        tracker_type = _resolve_choice(cell("Tracker Type"), TRACKER_TYPES) or default_tracker_type

        inserts.append((d, sys_num, family, issue_type, issue_desc, sps, ncr, status, solution, solution_date, crf, esw, scv, remarks, tracker_type))

    if not inserts:
        return 0, skipped

    with _db() as c:
        c.executemany(
            """
            INSERT INTO issues
            (date_reported, system_number, product_family, issue_type,
             issue_desc, sps_number, ncr_number, status,
             solution, solution_date, crf, esw, scv, remarks, tracker_type)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            """,
            inserts,
        )
        imported = len(inserts)

    return imported, skipped


def fetch_trend_data(tracker_scope="All", year_scope="All", month_scope="All", dimension="issue_type"):
    group_col = "issue_type" if dimension == "issue_type" else "product_family"
    params = []
    base_where = ["date_reported<>''"]
    if tracker_scope in TRACKER_TYPES:
        base_where.append("tracker_type=?")
        params.append(tracker_scope)

    if year_scope != "All":
        base_where.append("strftime('%Y', date_reported)=?")
        params.append(year_scope)

    if month_scope != "All":
        month_idx = datetime.datetime.strptime(month_scope, "%B").month
        base_where.append("CAST(strftime('%m', date_reported) AS INTEGER)=?")
        params.append(month_idx)

    where_sql = " AND ".join(base_where)
    with _db() as c:
        q = f"""
            SELECT CAST(strftime('%m', date_reported) AS INTEGER) AS month_idx,
                   {group_col} AS bucket,
                   COUNT(*) AS cnt
            FROM issues
            WHERE {where_sql}
            GROUP BY month_idx, bucket
        """
        rows = c.execute(q, params).fetchall()

    months = {i: defaultdict(int) for i in range(1, 13)}
    totals_by_bucket = defaultdict(int)
    if month_scope != "All":
        month_counts = defaultdict(int)
        for month_idx, bucket, cnt in rows:
            bucket = bucket or "Unspecified"
            month_counts[bucket] += cnt
            totals_by_bucket[bucket] += cnt
        return {
            "mode": "month",
            "dimension": dimension,
            "tracker_scope": tracker_scope,
            "year_scope": year_scope,
            "month_scope": month_scope,
            "buckets": month_counts,
            "totals": dict(totals_by_bucket),
        }

    for month_idx, bucket, cnt in rows:
        bucket = bucket or "Unspecified"
        months[int(month_idx or 0)][bucket] += cnt
        totals_by_bucket[bucket] += cnt

    return {
        "mode": "year",
        "dimension": dimension,
        "tracker_scope": tracker_scope,
        "year_scope": year_scope,
        "month_scope": month_scope,
        "months": {m: dict(months[m]) for m in range(1, 13)},
        "totals": dict(totals_by_bucket),
    }


def init_db():
    with _db() as c:
        c.execute(
            """
            CREATE TABLE IF NOT EXISTS issues (
                id             INTEGER PRIMARY KEY AUTOINCREMENT,
                date_reported  TEXT DEFAULT '',
                system_number  TEXT DEFAULT '',
                product_family TEXT DEFAULT '',
                issue_type     TEXT DEFAULT '',
                issue_desc     TEXT DEFAULT '',
                sps_number     TEXT DEFAULT '',
                ncr_number     TEXT DEFAULT '',
                status         TEXT DEFAULT 'Open',
                solution       TEXT DEFAULT '',
                solution_date  TEXT DEFAULT '',
                crf            TEXT DEFAULT '',
                esw            TEXT DEFAULT '',
                scv            TEXT DEFAULT '',
                remarks        TEXT DEFAULT '',
                tracker_type   TEXT DEFAULT 'Engineering',
                created_at     TEXT DEFAULT (datetime('now','localtime'))
            )
            """
        )
        existing = {r[1] for r in c.execute("PRAGMA table_info(issues)")}
        for col, defn in [
            ("ncr_number", "TEXT DEFAULT ''"),
            ("solution", "TEXT DEFAULT ''"),
            ("solution_date", "TEXT DEFAULT ''"),
            ("crf", "TEXT DEFAULT ''"),
            ("esw", "TEXT DEFAULT ''"),
            ("scv", "TEXT DEFAULT ''"),
            ("remarks", "TEXT DEFAULT ''"),
            ("tracker_type", "TEXT DEFAULT 'Engineering'"),
        ]:
            if col not in existing:
                c.execute(f"ALTER TABLE issues ADD COLUMN {col} {defn}")
        c.execute("CREATE INDEX IF NOT EXISTS idx_tracker_type ON issues(tracker_type)")


def fetch_issues(
    status_f="All",
    family_f="All",
    itype_f="All",
    month_f="All",
    year_f="All",
    tracker_type="Engineering",
):
    q = """
        SELECT id, date_reported, system_number, product_family, issue_type,
               issue_desc, sps_number, ncr_number, status,
               solution, solution_date, crf, esw, scv, remarks, tracker_type, created_at
        FROM issues WHERE 1=1
    """
    args = []
    if status_f != "All":
        q += " AND status=?"
        args.append(status_f)
    if family_f != "All":
        q += " AND product_family=?"
        args.append(family_f)
    if itype_f != "All":
        q += " AND issue_type=?"
        args.append(itype_f)
    if month_f != "All":
        q += " AND CAST(strftime('%m', date_reported) AS INTEGER)=?"
        args.append(datetime.datetime.strptime(month_f, "%B").month)
    if year_f != "All":
        q += " AND strftime('%Y', date_reported)=?"
        args.append(year_f)
    q += " AND tracker_type=?"
    args.append(tracker_type)
    q += " ORDER BY date_reported DESC, id DESC"
    with _db() as c:
        c.row_factory = _row_factory
        return c.execute(q, args).fetchall()


def fetch_by_id(db_id):
    with _db() as c:
        c.row_factory = _row_factory
        return c.execute("SELECT * FROM issues WHERE id=?", (db_id,)).fetchone()


def insert_issue(date_reported, system_number, product_family, issue_type, issue_desc, sps_number, ncr_number, tracker_type="Engineering"):
    with _db() as c:
        cur = c.execute(
            """
            INSERT INTO issues
            (date_reported, system_number, product_family, issue_type,
             issue_desc, sps_number, ncr_number, tracker_type)
            VALUES (?,?,?,?,?,?,?,?)
            """,
            (date_reported, system_number, product_family, issue_type, issue_desc, sps_number, ncr_number, tracker_type),
        )
        return cur.lastrowid


def update_issue(
    db_id,
    date_reported,
    system_number,
    product_family,
    issue_type,
    issue_desc,
    sps_number,
    ncr_number,
    status,
    solution,
    solution_date,
    crf,
    esw,
    scv,
    remarks,
    tracker_type,
):
    with _db() as c:
        c.execute(
            """
            UPDATE issues SET
                date_reported=?, system_number=?, product_family=?, issue_type=?,
                issue_desc=?, sps_number=?, ncr_number=?, status=?,
                solution=?, solution_date=?, crf=?, esw=?, scv=?, remarks=?, tracker_type=?
            WHERE id=?
            """,
            (
                date_reported,
                system_number,
                product_family,
                issue_type,
                issue_desc,
                sps_number,
                ncr_number,
                status,
                solution,
                solution_date,
                crf,
                esw,
                scv,
                remarks,
                tracker_type,
                db_id,
            ),
        )


def delete_by_ids(ids):
    with _db() as c:
        c.executemany("DELETE FROM issues WHERE id=?", [(i,) for i in ids])


def get_counts(tracker_type="Engineering"):
    with _db() as c:
        total = c.execute("SELECT COUNT(*) FROM issues WHERE tracker_type=?", (tracker_type,)).fetchone()[0]
        open_ = c.execute("SELECT COUNT(*) FROM issues WHERE status='Open' AND tracker_type=?", (tracker_type,)).fetchone()[0]
        closed = c.execute("SELECT COUNT(*) FROM issues WHERE status='Closed' AND tracker_type=?", (tracker_type,)).fetchone()[0]
        clarif = c.execute("SELECT COUNT(*) FROM issues WHERE status='Clarification' AND tracker_type=?", (tracker_type,)).fetchone()[0]
    return total, open_, closed, clarif


# -----------------------------------------------------------------------------
#  EXCEL EXPORT
# -----------------------------------------------------------------------------

def _fmt_date(d):
    try:
        return datetime.datetime.strptime(d, "%Y-%m-%d").strftime("%d %b %Y")
    except Exception:
        return d or ""


def _xl_border():
    s = Side(style="thin", color="B0C8E0")
    return Border(left=s, right=s, top=s, bottom=s)


def _auto_row_height(text, col_width_chars, base_pt=13, min_pt=18, padding=4):
    if not text:
        return min_pt
    lines = sum(max(1, -(-len(p) // max(1, col_width_chars))) for p in str(text).splitlines())
    return max(min_pt, lines * base_pt + padding)


def export_excel(engineering_rows, material_rows, filepath, customer_mode=False, export_scope="Both"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Production Issues"

    if customer_mode:
        headers = ["Date Reported", "System Number", "Product Family", "Issue Type", "Issue Description", "SPS Number", "Remarks"]
        col_widths = [14, 22, 14, 18, 60, 13, 35]
        left_cols = {4, 6}
    else:
        headers = [
            "Date Reported",
            "System Number",
            "Product Family",
            "Issue Type",
            "Issue Description",
            "SPS Number",
            "NCR Number",
            "Status",
            "Solution",
            "Solution Date",
            "CRF",
            "ESW",
            "SCV",
            "Remarks",
        ]
        col_widths = [14, 22, 14, 18, 55, 13, 13, 10, 45, 14, 12, 12, 12, 35]
        left_cols = {4, 8, 13}

    hdr_fill = PatternFill("solid", fgColor="1B5DA8")
    hdr_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    title_font = Font(name="Calibri", bold=True, color="FFFFFF", size=16)
    ctr = Alignment(horizontal="center", vertical="top", wrap_text=True)
    lft = Alignment(horizontal="left", vertical="top", wrap_text=True)
    title_align = Alignment(horizontal="center", vertical="center")

    current_row = 1

    def write_block(rows, title, start_row):
        if not rows:
            return start_row

        ws.merge_cells(f"A{start_row}:{get_column_letter(len(headers))}{start_row}")
        cell = ws.cell(row=start_row, column=1, value=title)
        cell.font = title_font
        cell.fill = PatternFill("solid", fgColor="007EA5")
        cell.alignment = title_align

        start_row += 1
        for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
            cell = ws.cell(row=start_row, column=ci, value=h)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = ctr
            cell.border = _xl_border()
            ws.column_dimensions[get_column_letter(ci)].width = w
        ws.row_dimensions[start_row].height = 26
        start_row += 1

        open_fill = PatternFill("solid", fgColor="FDEDEC")
        closed_fill = PatternFill("solid", fgColor="EAFAF1")
        alt_fill = PatternFill("solid", fgColor="EBF4FF")

        for row in rows:
            fill = open_fill if row["status"] == "Open" else closed_fill if row["status"] == "Closed" else alt_fill
            values = (
                [_fmt_date(row["date_reported"]), row["system_number"], row["product_family"], row["issue_type"], row["issue_desc"], row["sps_number"], row["remarks"]]
                if customer_mode
                else [
                    _fmt_date(row["date_reported"]),
                    row["system_number"],
                    row["product_family"],
                    row["issue_type"],
                    row["issue_desc"],
                    row["sps_number"],
                    row["ncr_number"],
                    row["status"],
                    row["solution"],
                    _fmt_date(row["solution_date"]),
                    row["crf"],
                    row["esw"],
                    row["scv"],
                    row["remarks"],
                ]
            )
            row_h = max((_auto_row_height(val, cw) for ci0, (val, cw) in enumerate(zip(values, col_widths)) if ci0 in left_cols), default=18)
            ws.row_dimensions[start_row].height = row_h
            for ci, (val, cw) in enumerate(zip(values, col_widths), 1):
                cell = ws.cell(row=start_row, column=ci, value=val)
                cell.fill = fill
                cell.border = _xl_border()
                cell.font = Font(name="Calibri", size=10)
                cell.alignment = lft if (ci - 1) in left_cols else ctr
            start_row += 1
        return start_row

    if export_scope == "Engineering":
        current_row = write_block(engineering_rows, "ENGINEERING ISSUES LOG", current_row)
    elif export_scope == "Material":
        current_row = write_block(material_rows, "MATERIAL ISSUES LOG", current_row)
    else:
        current_row = write_block(engineering_rows, "ENGINEERING ISSUES LOG", current_row)
        current_row += 2
        current_row = write_block(material_rows, "MATERIAL ISSUES LOG", current_row)

    wb.save(filepath)


# -----------------------------------------------------------------------------
#  UI HELPERS
# -----------------------------------------------------------------------------

def _fmt_display_date(value):
    if not value:
        return ""
    try:
        return datetime.datetime.strptime(value, "%Y-%m-%d").strftime("%d %b %Y")
    except Exception:
        return value


def _parse_iso_date(value):
    if not value:
        return None
    try:
        return datetime.datetime.strptime(value, "%Y-%m-%d").date()
    except Exception:
        return None


def _stylesheet():
    return f"""
    QMainWindow, QWidget#AppRoot {{
        background: {C['bg']};
        color: {C['text']};
        font-family: '{F['family']}';
        font-size: {F['size_md']}pt;
    }}
    QLabel, QRadioButton, QCheckBox {{
        background: transparent;
        color: {C['text']};
    }}
    QFrame#HeaderBar {{
        background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                                    stop:0 {C['header']},
                                    stop:0.55 #111c33,
                                    stop:1 #19294a);
        border: none;
        border-radius: 18px;
    }}
    QLabel#TitleLabel {{
        color: white;
        font-size: {F['size_xl']}pt;
        font-weight: 700;
    }}
    QLabel#MutedLabel {{
        color: {C['subtle']};
    }}
    QPushButton {{
        background: {C['accent']};
        color: #101828;
        border: 1px solid transparent;
        padding: 8px 14px;
        border-radius: 10px;
        font-weight: 700;
    }}
    QPushButton:hover {{
        background: #ffbf2f;
        color: #101828;
    }}
    QPushButton:pressed {{
        background: {C['accent_h']};
        color: white;
    }}
    QPushButton[outline="true"] {{
        background: {C['surface']};
        color: {C['text']};
        border: 1px solid {C['accent']};
    }}
    QPushButton[outline="true"]:hover {{
        background: {C['sel']};
        color: white;
    }}
    QPushButton[outline="true"]:pressed {{
        background: {C['accent']};
        color: #101828;
    }}
    QPushButton[danger="true"] {{
        background: #a02020;
        color: white;
    }}
    QPushButton[danger="true"]:hover {{
        background: #c62828;
        color: white;
    }}
    QPushButton[danger="true"]:pressed {{
        background: #7f1717;
        color: white;
    }}
    QToolButton {{
        background: {C['accent']};
        color: #101828;
        border: 1px solid transparent;
        border-radius: 10px;
        padding: 0px;
    }}
    QToolButton:hover {{
        background: #ffbf2f;
        color: #101828;
    }}
    QToolButton:pressed {{
        background: {C['accent_h']};
        color: white;
    }}
    QLineEdit, QComboBox {{
        background: {C['entry_bg']};
        border: 1px solid {C['border']};
        border-radius: 10px;
        padding: 6px;
        padding-right: 22px;
        selection-background-color: {C['sel']};
        selection-color: {C['text']};
    }}
    QPlainTextEdit, QTextEdit {{
        background: {C['entry_bg']};
        border: 1px solid {C['border']};
        border-radius: 10px;
        padding: 6px;
        padding-right: 6px;
        selection-background-color: {C['sel']};
        selection-color: {C['text']};
    }}
    QComboBox::drop-down {{
        subcontrol-origin: padding;
        subcontrol-position: top right;
        border: none;
        width: 18px;
        background: transparent;
    }}
    QPlainTextEdit QScrollBar:vertical, QTextEdit QScrollBar:vertical {{
        background: transparent;
        width: 10px;
        margin: 0px;
    }}
    QPlainTextEdit QScrollBar::handle:vertical, QTextEdit QScrollBar::handle:vertical {{
        background: {C['border']};
        min-height: 24px;
        border-radius: 5px;
    }}
    QPlainTextEdit QScrollBar::handle:vertical:hover, QTextEdit QScrollBar::handle:vertical:hover {{
        background: {C['accent']};
    }}
    QPlainTextEdit QScrollBar::add-line:vertical, QPlainTextEdit QScrollBar::sub-line:vertical,
    QTextEdit QScrollBar::add-line:vertical, QTextEdit QScrollBar::sub-line:vertical {{
        background: transparent;
        border: none;
        height: 0px;
    }}
    QPlainTextEdit QScrollBar::add-page:vertical, QPlainTextEdit QScrollBar::sub-page:vertical,
    QTextEdit QScrollBar::add-page:vertical, QTextEdit QScrollBar::sub-page:vertical {{
        background: transparent;
    }}
    QTabWidget::pane {{
        border: 0;
        margin-top: 0px;
    }}
    QTabWidget {{
        border: 0;
        background: transparent;
    }}
    QTabBar {{
        border-bottom: 1px solid {C['border']};
    }}
    QTabBar::tab {{
        background: {C['surface']};
        color: {C['subtle']};
        padding: 11px 20px;
        margin-right: 4px;
        margin-bottom: 0px;
        border-top-left-radius: 12px;
        border-top-right-radius: 12px;
        border: 1px solid {C['border']};
    }}
    QTabBar::tab:selected {{
        background: {C['panel']};
        color: {C['text']};
        border-bottom-color: {C['panel']};
    }}
    QTableWidget {{
        background: {C['surface']};
        alternate-background-color: {C['stripe']};
        gridline-color: {C['border']};
        border: 1px solid {C['border']};
        selection-background-color: {C['sel']};
        selection-color: {C['text']};
        border-radius: 12px;
    }}
    QHeaderView::section {{
        background: {C['surface_2']};
        color: white;
        padding: 6px;
        border: 1px solid {C['border']};
        font-weight: 700;
    }}
    QRadioButton {{
        spacing: 8px;
        color: {C['text']};
    }}
    QScrollArea {{
        border: none;
        background: transparent;
    }}
    QScrollBar:vertical {{
        background: transparent;
        width: 12px;
        margin: 2px 2px 2px 2px;
    }}
    QScrollBar::handle:vertical {{
        background: {C['border']};
        min-height: 30px;
        border-radius: 6px;
    }}
    QScrollBar::handle:vertical:hover {{
        background: {C['accent']};
    }}
    QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {{
        background: transparent;
        border: none;
        height: 0px;
    }}
    QScrollBar::add-page:vertical, QScrollBar::sub-page:vertical {{
        background: transparent;
    }}
    QScrollBar:horizontal {{
        background: transparent;
        height: 12px;
        margin: 2px 2px 2px 2px;
    }}
    QScrollBar::handle:horizontal {{
        background: {C['border']};
        min-width: 30px;
        border-radius: 6px;
    }}
    QScrollBar::handle:horizontal:hover {{
        background: {C['accent']};
    }}
    QScrollBar::add-line:horizontal, QScrollBar::sub-line:horizontal {{
        background: transparent;
        border: none;
        width: 0px;
    }}
    QScrollBar::add-page:horizontal, QScrollBar::sub-page:horizontal {{
        background: transparent;
    }}
    """


def _mk_label(text="", bold=False, color=None, object_name=None):
    label = QLabel(text)
    if object_name:
        label.setObjectName(object_name)
    if color:
        label.setStyleSheet(f"color: {color};")
    font = QFont(F["family"], F["size_md"])
    font.setBold(bold)
    label.setFont(font)
    return label


def _mk_button(text, on_click, outline=False, danger=False, width=None, color=None, text_color=None):
    btn = QPushButton(text)
    btn.setProperty("outline", outline)
    btn.setProperty("danger", danger)
    if width:
        btn.setMinimumWidth(width)
    btn.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
    if color and not outline and not danger:
        btn.setStyleSheet(f"background: {color}; color: {text_color or 'white'};")
    if on_click:
        btn.clicked.connect(on_click)
    btn.style().unpolish(btn)
    btn.style().polish(btn)
    return btn


def _set_field_width(widget, width=520):
    widget.setMinimumWidth(max(140, int(width * 0.33)))
    widget.setSizePolicy(QSizePolicy.Policy.Expanding, widget.sizePolicy().verticalPolicy())
    return widget


def _draw_icon(kind, color):
    pix = QPixmap(16, 16)
    pix.fill(Qt.GlobalColor.transparent)
    p = QPainter(pix)
    p.setRenderHint(QPainter.RenderHint.Antialiasing, True)
    pen = QPen(QColor(color))
    pen.setWidth(1)
    p.setPen(pen)
    if kind == "calendar":
        p.drawRoundedRect(2, 3, 12, 11, 2, 2)
        p.fillRect(2, 5, 12, 2, QColor(color))
        p.drawLine(5, 1, 5, 5)
        p.drawLine(11, 1, 11, 5)
        p.drawPoint(6, 8)
        p.drawPoint(10, 8)
    elif kind == "clear":
        pen.setWidth(2)
        p.setPen(pen)
        p.drawLine(4, 4, 12, 12)
        p.drawLine(12, 4, 4, 12)
    elif kind == "arrow_left":
        pen.setWidth(2)
        p.setPen(pen)
        p.drawLine(10, 3, 6, 8)
        p.drawLine(6, 8, 10, 13)
    elif kind == "arrow_right":
        pen.setWidth(2)
        p.setPen(pen)
        p.drawLine(6, 3, 10, 8)
        p.drawLine(10, 8, 6, 13)
    p.end()
    return QIcon(pix)


def _card(title=None):
    frame = QFrame()
    frame.setObjectName("Card")
    frame.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Preferred)
    frame.setStyleSheet(
        f"""
        QFrame#Card {{
            background: {C['surface']};
            border: 1px solid {C['border']};
            border-radius: 14px;
        }}
        """
    )
    if title is None:
        return frame
    layout = QVBoxLayout(frame)
    layout.setContentsMargins(14, 14, 14, 14)
    layout.setSpacing(10)
    layout.addWidget(_mk_label(title, bold=True, color=C["subtle"]))
    return frame


class ClickableCard(QFrame):
    clicked = Signal()

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setCursor(Qt.PointingHandCursor)
        self.setAttribute(Qt.WidgetAttribute.WA_Hover, True)
        self._selected = False

    def mousePressEvent(self, event):
        if event.button() == Qt.MouseButton.LeftButton:
            self.clicked.emit()
        super().mousePressEvent(event)


class FilterDropdown(QFrame):
    def __init__(self, items, parent=None):
        super().__init__(parent)
        self.setObjectName("FilterDropdown")
        self.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        self._build(items)

    def _build(self, items):
        lay = QHBoxLayout(self)
        lay.setContentsMargins(0, 0, 0, 0)
        lay.setSpacing(0)
        self.setStyleSheet(
            f"""
            QFrame#FilterDropdown {{
                background: {C['entry_bg']};
                border: 1px solid {C['border']};
                border-radius: 10px;
            }}
            """
        )

        self.combo = QComboBox(self)
        self.combo.addItems(items)
        self.combo.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        self.combo.setStyleSheet(self.combo.styleSheet() + """
            QComboBox {
                background: transparent;
                border: none;
                padding: 6px 8px 6px 8px;
            }
        """)
        lay.addWidget(self.combo, 1)
        arrow = QLabel("▾", self)
        arrow.setAlignment(Qt.AlignCenter)
        arrow.setFixedWidth(18)
        arrow.setStyleSheet(f"color: {C['subtle']}; background: transparent;")
        lay.addWidget(arrow)

        self.currentTextChanged = self.combo.currentTextChanged

    def currentText(self):
        return self.combo.currentText()

    def setCurrentText(self, text):
        self.combo.setCurrentText(text)

    def blockSignals(self, block):
        return self.combo.blockSignals(block)

    def addItems(self, items):
        self.combo.addItems(items)

    def mousePressEvent(self, event):
        if event.button() == Qt.MouseButton.LeftButton:
            self.combo.showPopup()
        super().mousePressEvent(event)


def _metric_card(title, value="", accent=None):
    card = ClickableCard()
    card.setObjectName("MetricCard")
    card.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Preferred)
    card.setStyleSheet(
        f"""
        QFrame#MetricCard {{
            background: {C['surface']};
            border: 1px solid {C['border']};
            border-radius: 14px;
        }}
        QFrame#MetricCard:hover {{
            background: {C['sel']};
            border: 1px solid {C['accent']};
        }}
        QFrame#MetricCard[selected="true"] {{
            background: {C['sel']};
            border: 1px solid {C['accent']};
        }}
        """
    )
    lay = QVBoxLayout(card)
    lay.setContentsMargins(16, 16, 16, 16)
    lay.setSpacing(4)
    title_lbl = _mk_label(title, bold=True, color=C["subtle"])
    value_lbl = _mk_label(value, bold=True)
    value_lbl.setStyleSheet(f"font-size: 18pt; color: {accent or C['text']};")
    lay.addWidget(title_lbl)
    lay.addWidget(value_lbl)
    lay.addStretch(1)
    card._value_label = value_lbl
    return card


class CalendarDialog(QDialog):
    def __init__(self, parent=None, initial=None):
        super().__init__(parent)
        self.setWindowTitle("Select Date")
        self.setWindowFlags(Qt.Tool | Qt.FramelessWindowHint)
        self.setAttribute(Qt.WidgetAttribute.WA_DeleteOnClose, True)
        self.result_date = None
        self._build(initial)

    def _build(self, initial):
        self.setStyleSheet(_stylesheet())
        lay = QVBoxLayout(self)
        lay.setContentsMargins(10, 10, 10, 10)
        lay.setSpacing(10)

        self.calendar = QCalendarWidget(self)
        self.calendar.setGridVisible(True)
        self.calendar.setVerticalHeaderFormat(QCalendarWidget.VerticalHeaderFormat.NoVerticalHeader)
        self.calendar.setNavigationBarVisible(False)
        self.calendar.setStyleSheet(
            f"""
            QCalendarWidget {{
                background: {C['surface']};
                color: {C['text']};
                border: 1px solid {C['border']};
                border-radius: 12px;
            }}
            QCalendarWidget QWidget {{
                alternate-background-color: {C['surface']};
            }}
            QCalendarWidget QMenu {{
                background: {C['surface']};
                color: {C['text']};
                border: 1px solid {C['border']};
            }}
            QCalendarWidget QSpinBox {{
                background: {C['entry_bg']};
                color: {C['text']};
                border: 1px solid {C['border']};
                border-radius: 8px;
                padding: 2px 6px;
            }}
            QCalendarWidget QAbstractItemView {{
                selection-background-color: {C['accent']};
                selection-color: #101828;
                background: {C['surface']};
                color: {C['text']};
                border: none;
                outline: 0;
            }}
            QCalendarWidget QAbstractItemView:enabled {{
                gridline-color: {C['border']};
            }}
            QCalendarWidget QAbstractItemView::item:selected {{
                background: {C['accent']};
                color: #101828;
            }}
            QCalendarWidget QAbstractItemView::item:hover {{
                background: {C['sel']};
            }}
            QCalendarWidget QAbstractItemView:disabled {{
                color: {C['subtle']};
            }}
            """
        )
        weekend_fmt = QTextCharFormat()
        weekend_fmt.setForeground(QColor(C["accent"]))
        self.calendar.setWeekdayTextFormat(Qt.DayOfWeek.Saturday, weekend_fmt)
        self.calendar.setWeekdayTextFormat(Qt.DayOfWeek.Sunday, weekend_fmt)

        today = datetime.date.today()
        current_year = today.year
        initial_date = None
        if initial:
            initial_date = QDate(initial.year, initial.month, initial.day)
        else:
            initial_date = QDate(today.year, today.month, today.day)
        self.calendar.setSelectedDate(initial_date)
        self.calendar.setCurrentPage(initial_date.year(), initial_date.month())

        header = QFrame(self)
        header.setObjectName("CalendarHeader")
        header.setStyleSheet(
            f"""
            QFrame#CalendarHeader {{
                background: {C['surface_2']};
                border: 1px solid {C['border']};
                border-radius: 12px;
            }}
            """
        )
        header_lay = QHBoxLayout(header)
        header_lay.setContentsMargins(10, 8, 10, 8)
        header_lay.setSpacing(8)

        def nav_btn(text, on_click):
            btn = QToolButton(self)
            btn.setText(text)
            btn.setFixedSize(28, 28)
            btn.setCursor(Qt.PointingHandCursor)
            btn.clicked.connect(on_click)
            btn.setStyleSheet(
                f"""
                QToolButton {{
                    background: {C['accent']};
                    color: #101828;
                    border: none;
                    border-radius: 9px;
                    font-size: 13pt;
                    font-weight: 700;
                }}
                QToolButton:hover {{
                    background: {C['accent_h']};
                    color: white;
                }}
                """
            )
            return btn

        self.prev_btn = nav_btn("‹", self._prev_month)
        self.next_btn = nav_btn("›", self._next_month)

        self.month_combo = QComboBox(self)
        self.month_combo.addItems([calendar.month_name[i] for i in range(1, 13)])
        self.month_combo.setCurrentIndex(initial_date.month() - 1)
        self.month_combo.currentIndexChanged.connect(self._jump_to_page)

        self.year_combo = QComboBox(self)
        self.year_combo.addItems([str(y) for y in range(current_year - 20, current_year + 21)])
        year_index = self.year_combo.findText(str(initial_date.year()))
        if year_index >= 0:
            self.year_combo.setCurrentIndex(year_index)
        self.year_combo.currentIndexChanged.connect(self._jump_to_page)

        for btn, kind in ((self.prev_btn, "arrow_left"), (self.next_btn, "arrow_right")):
            btn.setText("")
            btn.setIcon(_draw_icon(kind, "#101828"))
            btn.setIconSize(QSize(14, 14))
            btn.setFixedSize(30, 30)

        self.month_combo.setMinimumWidth(120)
        self.year_combo.setMinimumWidth(92)
        self.month_combo.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        self.year_combo.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        self.month_combo.setCursor(Qt.PointingHandCursor)
        self.year_combo.setCursor(Qt.PointingHandCursor)

        month_wrap = QWidget(self)
        month_lay = QHBoxLayout(month_wrap)
        month_lay.setContentsMargins(0, 0, 0, 0)
        month_lay.setSpacing(6)
        month_lay.addWidget(self.month_combo)
        month_lay.addWidget(self.year_combo)

        header_lay.addWidget(self.prev_btn)
        header_lay.addWidget(month_wrap)
        header_lay.addStretch(1)
        header_lay.addWidget(self.next_btn)

        self.calendar.currentPageChanged.connect(self._sync_header)
        self._sync_header(initial_date.year(), initial_date.month())

        lay.addWidget(header)
        self.calendar.clicked.connect(self._choose)
        lay.addWidget(self.calendar)

        row = QHBoxLayout()
        row.addStretch(1)
        row.addWidget(_mk_button("Today", self._pick_today, outline=True, width=80))
        row.addWidget(_mk_button("Cancel", self.reject, outline=True, width=80))
        lay.addLayout(row)

    def _choose(self, qdate):
        self.result_date = qdate.toPython().strftime("%Y-%m-%d")
        self.accept()

    def _pick_today(self):
        self.result_date = datetime.date.today().strftime("%Y-%m-%d")
        self.accept()

    def _sync_header(self, year, month):
        with QSignalBlocker(self.month_combo), QSignalBlocker(self.year_combo):
            self.month_combo.setCurrentIndex(month - 1)
            year_index = self.year_combo.findText(str(year))
            if year_index >= 0:
                self.year_combo.setCurrentIndex(year_index)

    def _jump_to_page(self):
        self.calendar.setCurrentPage(int(self.year_combo.currentText()), self.month_combo.currentIndex() + 1)

    def _prev_month(self):
        year = self.calendar.yearShown()
        month = self.calendar.monthShown() - 1
        if month < 1:
            month = 12
            year -= 1
        self.calendar.setCurrentPage(year, month)

    def _next_month(self):
        year = self.calendar.yearShown()
        month = self.calendar.monthShown() + 1
        if month > 12:
            month = 1
            year += 1
        self.calendar.setCurrentPage(year, month)

    def _jump_to_page(self, *_):
        self.calendar.setCurrentPage(int(self.year_combo.currentText()), self.month_combo.currentIndex() + 1)


class DateFieldWidget(QWidget):
    def __init__(self, parent=None, initial=None, nullable=False):
        super().__init__(parent)
        self._nullable = nullable
        self._value = initial or ""
        self._calendar_dialog = None
        if not self._value and not nullable:
            self._value = datetime.date.today().strftime("%Y-%m-%d")
        self._build()
        self.set_value(self._value)

    def _build(self):
        lay = QHBoxLayout(self)
        lay.setContentsMargins(0, 0, 0, 0)
        lay.setSpacing(4)

        self.edit = QLineEdit(self)
        self.edit.setReadOnly(True)
        self.edit.setMinimumWidth(0)
        self.edit.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        lay.addWidget(self.edit)

        self.btn_cal = QToolButton(self)
        self.btn_cal.setIcon(_draw_icon("calendar", "#ffffff"))
        self.btn_cal.setIconSize(QSize(16, 16))
        self.btn_cal.setFixedSize(30, 30)
        self.btn_cal.clicked.connect(self.open_calendar)
        lay.addWidget(self.btn_cal)

        if self._nullable:
            self.btn_clear = QToolButton(self)
            self.btn_clear.setIcon(_draw_icon("clear", "#ffffff"))
            self.btn_clear.setIconSize(QSize(14, 14))
            self.btn_clear.setFixedSize(30, 30)
            self.btn_clear.clicked.connect(self.clear_value)
            lay.addWidget(self.btn_clear)
        else:
            self.btn_clear = None

    def set_value(self, value):
        self._value = value or ""
        self.edit.setText(_fmt_display_date(self._value))

    def clear_value(self):
        if self._nullable:
            self.set_value("")

    def open_calendar(self):
        if self._calendar_dialog is not None and self._calendar_dialog.isVisible():
            self._calendar_dialog.close()
            return
        initial = _parse_iso_date(self._value)
        dlg = CalendarDialog(self, initial=initial)
        self._calendar_dialog = dlg
        dlg.adjustSize()
        anchor = self.btn_cal.mapToGlobal(QPoint(0, self.btn_cal.height()))
        screen = QGuiApplication.screenAt(anchor) or self.btn_cal.screen() or QGuiApplication.primaryScreen()
        available = screen.availableGeometry() if screen else None

        x = anchor.x() - dlg.width() + self.btn_cal.width()
        below_y = anchor.y() + 6
        above_y = self.btn_cal.mapToGlobal(QPoint(0, 0)).y() - dlg.height() - 6

        if available is not None:
            if below_y + dlg.height() > available.bottom() and above_y >= available.top():
                y = above_y
            else:
                y = below_y
            x = max(available.left(), min(x, available.right() - dlg.width()))
            y = max(available.top(), min(y, available.bottom() - dlg.height()))
        else:
            y = below_y

        dlg.move(x, y)
        dlg.accepted.connect(self._calendar_accepted)
        dlg.rejected.connect(self._calendar_closed)
        dlg.destroyed.connect(lambda *_: self._clear_calendar_dialog())
        dlg.show()
        dlg.raise_()
        dlg.activateWindow()

    def _calendar_accepted(self):
        dlg = self._calendar_dialog
        if dlg is not None and dlg.result_date:
            self.set_value(dlg.result_date)
        self._clear_calendar_dialog()

    def _calendar_closed(self):
        self._clear_calendar_dialog()

    def _clear_calendar_dialog(self):
        self._calendar_dialog = None

    def get(self):
        return self._value

    def set(self, value):
        self.set_value(value)


class CopyPasteDialog(QDialog):
    def __init__(self, parent, system_number, issue_desc, sps_number):
        super().__init__(parent)
        self.setWindowTitle("Copy-Paste Text")
        self.setModal(True)
        self.resize(640, 520)
        self._build(system_number, issue_desc, sps_number)

    def _build(self, sys_num, desc, sps_num):
        self.setStyleSheet(_stylesheet())
        systems_block = sys_num.strip()
        email_text = (
            f"Hi KB / Zach,\n\n"
            f"SPS {sps_num} submitted for the following issue(s):\n\n"
            f"{systems_block}\n\n"
            f"{desc}\n\n"
            f"Best regards,"
        )
        desc_of_def = f"{systems_block}\n\n{desc}"
        det_dispos = f"AMAT SPS {sps_num} submitted."

        outer = QVBoxLayout(self)
        outer.setContentsMargins(14, 14, 14, 14)
        outer.setSpacing(12)

        def block(title, content, h):
            outer.addWidget(_mk_label(title, bold=True, color=C["text"]))
            tb = QPlainTextEdit(self)
            tb.setPlainText(content)
            tb.setReadOnly(True)
            tb.setMinimumHeight(h)
            tb.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
            tb.setStyleSheet(
                f"""
                QPlainTextEdit {{
                    background: {C['entry_bg']};
                    color: {C['text']};
                    border: 1px solid {C['border']};
                    border-radius: 12px;
                    padding: 8px;
                }}
                """
            )
            outer.addWidget(tb)
            row = QHBoxLayout()
            row.addStretch(1)
            btn = _mk_button("Copy", lambda: self._copy_text(tb.toPlainText(), btn), outline=True, width=80)
            row.addWidget(btn)
            outer.addLayout(row)

        block("Email body", email_text, 170)
        block("NCR - Desc. of Def. / Req. for Change", desc_of_def, 130)
        block("NCR - Det. Dispos. / Reas. for Change", det_dispos, 80)

    def _copy_text(self, text, btn):
        QApplication.clipboard().setText(text.rstrip())
        btn.setText("Copied")
        QTimer.singleShot(1200, lambda: btn.setText("Copy"))


class NewIssueDialog(QDialog):
    def __init__(self, parent, tracker_type="Engineering"):
        super().__init__(parent)
        self.tracker_type = tracker_type
        self.setWindowTitle(f"Log New {tracker_type} Issue")
        self.resize(620, 575)
        self.setMinimumWidth(560)
        self.setModal(True)
        self._build()

    def _build(self):
        self.setStyleSheet(_stylesheet())
        lay = QVBoxLayout(self)
        lay.setContentsMargins(22, 18, 22, 18)
        lay.setSpacing(6)
        field_w = 520

        lay.addWidget(_mk_label(f"New {self.tracker_type} Issue", bold=True))

        form = QGridLayout()
        form.setHorizontalSpacing(18)
        form.setVerticalSpacing(6)
        form.setColumnStretch(0, 0)
        form.setColumnStretch(1, 1)
        lay.addLayout(form)

        r = 0
        form.addWidget(_mk_label("Date Reported *"), r, 0, alignment=Qt.AlignLeft)
        self.dt = DateFieldWidget(self)
        _set_field_width(self.dt, field_w)
        form.addWidget(self.dt, r, 1)

        r += 1
        form.addWidget(_mk_label("System Number(s) *"), r, 0, alignment=Qt.AlignTop)
        sys_box = QVBoxLayout()
        sys_box.setContentsMargins(0, 0, 0, 0)
        sys_box.setSpacing(4)
        self.sys = QPlainTextEdit(self)
        self.sys.setMinimumHeight(60)
        self.sys.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.MinimumExpanding)
        _set_field_width(self.sys, field_w)
        sys_box.addWidget(self.sys)
        sys_hint = _mk_label("One system per line for multiple systems.", color=C["subtle"])
        sys_hint.setWordWrap(True)
        sys_box.addWidget(sys_hint)
        sys_wrap = QWidget()
        sys_wrap.setLayout(sys_box)
        sys_wrap.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        form.addWidget(sys_wrap, r, 1)

        r += 1
        form.addWidget(_mk_label("Product Family *"), r, 0)
        self.fam = QComboBox(self)
        self.fam.addItems(FAMILIES)
        _set_field_width(self.fam, field_w)
        form.addWidget(self.fam, r, 1)

        r += 1
        form.addWidget(_mk_label("Issue Type *"), r, 0)
        self.ityp = QComboBox(self)
        self.ityp.addItems(ISSUE_TYPES)
        default_issue_type = "Request for Deviation" if self.tracker_type == "Material" else ISSUE_TYPES[0]
        self.ityp.setCurrentText(default_issue_type)
        _set_field_width(self.ityp, field_w)
        form.addWidget(self.ityp, r, 1)

        r += 1
        form.addWidget(_mk_label("Issue Description *"), r, 0, alignment=Qt.AlignTop)
        desc_box = QVBoxLayout()
        desc_box.setContentsMargins(0, 0, 0, 0)
        desc_box.setSpacing(4)
        self.desc = QPlainTextEdit(self)
        self.desc.setMinimumHeight(130)
        self.desc.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        _set_field_width(self.desc, field_w)
        desc_box.addWidget(self.desc)
        desc_hint = _mk_label("Separate multiple issues with a blank line.", color=C["subtle"])
        desc_hint.setWordWrap(True)
        desc_box.addWidget(desc_hint)
        desc_wrap = QWidget()
        desc_wrap.setLayout(desc_box)
        desc_wrap.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Preferred)
        form.addWidget(desc_wrap, r, 1)

        r += 1
        form.addWidget(_mk_label("SPS Number *"), r, 0)
        self.sps = QLineEdit(self)
        self.sps.setPlaceholderText("e.g. 752766")
        _set_field_width(self.sps, field_w)
        form.addWidget(self.sps, r, 1)

        r += 1
        form.addWidget(_mk_label("NCR Number"), r, 0)
        self.ncr = QLineEdit(self)
        self.ncr.setPlaceholderText("e.g. NCR281174 (fill after creating in Agile)")
        _set_field_width(self.ncr, field_w)
        form.addWidget(self.ncr, r, 1)

        buttons = QHBoxLayout()
        buttons.addStretch(1)
        buttons.addWidget(_mk_button("Cancel", self.reject, outline=True, width=90))
        buttons.addWidget(_mk_button("Generate Copy-Paste Text", self._generate, color=C["subtle"], width=210))
        buttons.addWidget(_mk_button("Submit", self._submit, width=120))
        lay.addLayout(buttons)

    def _generate(self):
        dlg = CopyPasteDialog(self, self.sys.toPlainText().strip(), self.desc.toPlainText().strip(), self.sps.text().strip())
        dlg.exec()

    def _submit(self):
        d = self.dt.get().strip()
        s = self.sys.toPlainText().strip()
        desc = self.desc.toPlainText().strip()
        sps = self.sps.text().strip()
        if not all([d, s, desc, sps]):
            QMessageBox.critical(self, "Missing Fields", "Date Reported, System Number, Issue Description and SPS Number are required.")
            return
        insert_issue(d, s, self.fam.currentText(), self.ityp.currentText(), desc, sps, self.ncr.text().strip(), self.tracker_type)
        self.accept()


class ManageDialog(QDialog):
    def __init__(self, parent, db_id):
        super().__init__(parent)
        self.db_id = db_id
        self.R = fetch_by_id(db_id)
        self.setWindowTitle("Edit Issue")
        self.resize(720, 880)
        self.setModal(True)
        self._build()

    def _build(self):
        self.setStyleSheet(_stylesheet())
        outer = QVBoxLayout(self)
        outer.setContentsMargins(0, 0, 0, 0)
        outer.setSpacing(0)
        field_w = 520
        label_w = 125

        scroll = SmoothScrollArea(self)
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.NoFrame)
        scroll.verticalScrollBar().setSingleStep(8)
        scroll.horizontalScrollBar().setSingleStep(8)
        outer.addWidget(scroll)

        content = QWidget()
        scroll.setWidget(content)
        lay = QVBoxLayout(content)
        lay.setContentsMargins(22, 18, 22, 18)
        lay.setSpacing(10)

        R = self.R
        title_sys = R["system_number"].splitlines()[0] if R["system_number"] else ""
        tracker_type = R.get("tracker_type", "Engineering")
        lay.addWidget(_mk_label(f"{tracker_type} Issue - {title_sys}", bold=True))
        created = _mk_label(f"Created: {R['created_at']}", color=C["subtle"])
        lay.addWidget(created)

        form = QGridLayout()
        form.setContentsMargins(0, 0, 0, 0)
        form.setHorizontalSpacing(18)
        form.setVerticalSpacing(10)
        form.setColumnStretch(0, 0)
        form.setColumnStretch(1, 1)
        form.setColumnMinimumWidth(0, label_w)
        lay.addLayout(form)

        r = 0

        def section(title, row_idx):
            bar = QFrame(self)
            bar.setObjectName("HeaderBar")
            bar.setStyleSheet(f"background: {C['header']};")
            bar.setMinimumHeight(24)
            bar_row = QHBoxLayout(bar)
            bar_row.setContentsMargins(8, 2, 8, 2)
            bar_row.addWidget(_mk_label(title, bold=True, color="white"))
            form.addWidget(bar, row_idx, 0, 1, 2)
            return row_idx + 1

        r = section("Basic Information", r)

        form.addWidget(_mk_label("Date Reported"), r, 0)
        self.dt = DateFieldWidget(self, initial=R["date_reported"])
        _set_field_width(self.dt, field_w)
        form.addWidget(self.dt, r, 1)

        r += 1
        form.addWidget(_mk_label("System Number(s)"), r, 0, alignment=Qt.AlignTop)
        sys_box = QVBoxLayout()
        sys_box.setContentsMargins(0, 0, 0, 0)
        sys_box.setSpacing(4)
        self.sys = QPlainTextEdit(self)
        self.sys.setMinimumHeight(60)
        self.sys.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.MinimumExpanding)
        _set_field_width(self.sys, field_w)
        self.sys.setPlainText(R["system_number"] or "")
        sys_box.addWidget(self.sys)
        sys_hint = _mk_label("One system per line for multiple systems.", color=C["subtle"])
        sys_hint.setWordWrap(True)
        sys_box.addWidget(sys_hint)
        sys_wrap = QWidget()
        sys_wrap.setLayout(sys_box)
        sys_wrap.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        form.addWidget(sys_wrap, r, 1)

        r += 1
        form.addWidget(_mk_label("Product Family"), r, 0)
        self.fam = QComboBox(self)
        self.fam.addItems(FAMILIES)
        self.fam.setCurrentText(R["product_family"] or FAMILIES[0])
        _set_field_width(self.fam, field_w)
        form.addWidget(self.fam, r, 1)

        r += 1
        form.addWidget(_mk_label("Issue Type"), r, 0)
        self.ityp = QComboBox(self)
        self.ityp.addItems(ISSUE_TYPES)
        self.ityp.setCurrentText(R["issue_type"] or ISSUE_TYPES[0])
        _set_field_width(self.ityp, field_w)
        form.addWidget(self.ityp, r, 1)

        r += 1
        form.addWidget(_mk_label("Issue Description"), r, 0, alignment=Qt.AlignTop)
        desc_box = QVBoxLayout()
        desc_box.setContentsMargins(0, 0, 0, 0)
        desc_box.setSpacing(4)
        self.desc = QPlainTextEdit(self)
        self.desc.setMinimumHeight(100)
        self.desc.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        _set_field_width(self.desc, field_w)
        self.desc.setPlainText(R["issue_desc"] or "")
        desc_box.addWidget(self.desc)
        desc_hint = _mk_label("Separate multiple issues with a blank line.", color=C["subtle"])
        desc_hint.setWordWrap(True)
        desc_box.addWidget(desc_hint)
        desc_wrap = QWidget()
        desc_wrap.setLayout(desc_box)
        desc_wrap.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Preferred)
        form.addWidget(desc_wrap, r, 1)

        r += 1
        form.addWidget(_mk_label("SPS Number"), r, 0)
        self.sps = QLineEdit(self)
        self.sps.setText(R["sps_number"] or "")
        _set_field_width(self.sps, field_w)
        form.addWidget(self.sps, r, 1)

        r += 1
        form.addWidget(_mk_label("NCR Number"), r, 0)
        self.ncr = QLineEdit(self)
        self.ncr.setText(R["ncr_number"] or "")
        _set_field_width(self.ncr, field_w)
        form.addWidget(self.ncr, r, 1)

        r += 1
        form.addWidget(_mk_label("Status"), r, 0)
        status_row = QHBoxLayout()
        status_row.setContentsMargins(0, 0, 0, 0)
        status_row.setSpacing(16)
        self.open_rb = QRadioButton("Open")
        self.clarif_rb = QRadioButton("Clarification")
        self.closed_rb = QRadioButton("Closed")
        radio_styles = {
            self.open_rb: (C["open"], "Open"),
            self.clarif_rb: (C["clarif"], "Clarification"),
            self.closed_rb: (C["closed"], "Closed"),
        }
        for rb, (accent, _) in radio_styles.items():
            rb.setCursor(Qt.PointingHandCursor)
            rb.setStyleSheet(
                f"""
                QRadioButton {{
                    background: {C['surface']};
                    color: {accent};
                    border: 1px solid {C['border']};
                    border-radius: 10px;
                    padding: 8px 12px;
                    spacing: 8px;
                    font-weight: 700;
                    min-height: 18px;
                }}
                QRadioButton:hover {{
                    background: {C['sel']};
                    border: 1px solid {accent};
                }}
                QRadioButton:checked {{
                    background: {C['sel']};
                    border: 1px solid {accent};
                    color: white;
                }}
                QRadioButton::indicator {{
                    width: 12px;
                    height: 12px;
                }}
                QRadioButton::indicator:unchecked {{
                    border: 2px solid {C['subtle']};
                    border-radius: 6px;
                    background: transparent;
                }}
                QRadioButton::indicator:checked {{
                    border: 2px solid {accent};
                    border-radius: 6px;
                    background: {accent};
                }}
                """
            )
        for rb in (self.open_rb, self.clarif_rb, self.closed_rb):
            status_row.addWidget(rb)
            status_row.addSpacing(8)
        status_row.addStretch(1)
        current = R["status"] or "Open"
        if current == "Open":
            self.open_rb.setChecked(True)
        elif current == "Clarification":
            self.clarif_rb.setChecked(True)
        else:
            self.closed_rb.setChecked(True)
        status_wrap = QWidget()
        status_wrap.setLayout(status_row)
        status_wrap.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        _set_field_width(status_wrap, field_w)
        form.addWidget(status_wrap, r, 1)

        r += 1
        r = section("Resolution", r)

        form.addWidget(_mk_label("Solution"), r, 0, alignment=Qt.AlignTop)
        self.sol = QPlainTextEdit(self)
        self.sol.setMinimumHeight(70)
        self.sol.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        _set_field_width(self.sol, field_w)
        self.sol.setPlainText(R["solution"] or "")
        sol_wrap = QWidget()
        sol_lay = QVBoxLayout(sol_wrap)
        sol_lay.setContentsMargins(0, 0, 0, 0)
        sol_lay.setSpacing(0)
        sol_lay.addWidget(self.sol)
        sol_wrap.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Preferred)
        form.addWidget(sol_wrap, r, 1)

        r += 1
        form.addWidget(_mk_label("Solution Date"), r, 0)
        self.sol_dt = DateFieldWidget(self, initial=R["solution_date"] or None, nullable=True)
        _set_field_width(self.sol_dt, field_w)
        form.addWidget(self.sol_dt, r, 1)

        self.crf = QLineEdit(self)
        self.crf.setText(R["crf"] or "")
        self.esw = QLineEdit(self)
        self.esw.setText(R["esw"] or "")
        self.scv = QLineEdit(self)
        self.scv.setText(R["scv"] or "")
        r += 1
        r = section("Change Controls", r)
        for row_idx, (label, widget) in enumerate([("CRF", self.crf), ("ESW", self.esw), ("SCV", self.scv)]):
            form.addWidget(_mk_label(label), r + row_idx, 0)
            _set_field_width(widget, field_w)
            form.addWidget(widget, r + row_idx, 1)

        r += 3
        r = section("Remarks", r)
        form.addWidget(_mk_label("Remarks"), r, 0, alignment=Qt.AlignTop)
        self.rmk = QPlainTextEdit(self)
        self.rmk.setMinimumHeight(70)
        self.rmk.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        _set_field_width(self.rmk, field_w)
        self.rmk.setPlainText(R["remarks"] or "")
        rem_wrap = QWidget()
        rem_lay = QVBoxLayout(rem_wrap)
        rem_lay.setContentsMargins(0, 0, 0, 0)
        rem_lay.setSpacing(0)
        rem_lay.addWidget(self.rmk)
        rem_wrap.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Preferred)
        form.addWidget(rem_wrap, r, 1)

        lay.addStretch(1)

        footer = QFrame(self)
        footer.setStyleSheet(f"background: {C['panel']};")
        footer.setMinimumHeight(56)
        footer_row = QHBoxLayout(footer)
        footer_row.setContentsMargins(14, 10, 14, 10)
        footer_row.addWidget(_mk_button("Delete", self._delete, danger=True, width=100))
        footer_row.addStretch(1)
        footer_row.addWidget(_mk_button("Cancel", self.reject, outline=True, width=90))
        footer_row.addWidget(_mk_button("Save Changes", self._save, width=150, color="#ffffff", text_color="#101828"))
        outer.addWidget(footer)

    def _save(self):
        tracker_type = self.R.get("tracker_type", "Engineering")
        status = "Open"
        if self.clarif_rb.isChecked():
            status = "Clarification"
        elif self.closed_rb.isChecked():
            status = "Closed"
        update_issue(
            self.db_id,
            self.dt.get(),
            self.sys.toPlainText().strip(),
            self.fam.currentText(),
            self.ityp.currentText(),
            self.desc.toPlainText().strip(),
            self.sps.text().strip(),
            self.ncr.text().strip(),
            status,
            self.sol.toPlainText().strip(),
            self.sol_dt.get(),
            self.crf.text().strip(),
            self.esw.text().strip(),
            self.scv.text().strip(),
            self.rmk.toPlainText().strip(),
            tracker_type,
        )
        QMessageBox.information(self, "Saved", "Issue updated.")
        self.accept()

    def _delete(self):
        if QMessageBox.question(
            self,
            "Delete",
            "Permanently delete this issue?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No,
        ) == QMessageBox.StandardButton.Yes:
            delete_by_ids([self.db_id])
            self.accept()


class IssueTableWidget(QTableWidget):
    sortRequested = Signal(str)
    editRequested = Signal(int)
    deleteRequested = Signal(int)
    selectionChangedCount = Signal(int)

    def __init__(self, parent=None):
        super().__init__(parent)
        self._rows = []
        self._columns = TABLE_COLS
        self._col_index_by_id = {cid: idx + 1 for idx, (cid, *_rest) in enumerate(self._columns)}
        self.current_cell = (-1, -1)
        self._flash_cells = set()
        self._flash_border_color = QColor("#ffd76a")
        self._flash_timer = QTimer(self)
        self._flash_timer.setSingleShot(True)
        self._flash_timer.timeout.connect(self._clear_flash_cells)
        self._build()

    def _build(self):
        self.setColumnCount(1 + len(self._columns))
        self.setHorizontalHeaderLabels(["#"] + [c[1] for c in self._columns])
        self.setAlternatingRowColors(True)
        self.setWordWrap(True)
        self.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.setSelectionMode(QAbstractItemView.SelectionMode.ExtendedSelection)
        self.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.verticalHeader().setVisible(False)
        self.verticalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Fixed)
        self.horizontalHeader().setSectionsClickable(True)
        self.horizontalHeader().setSortIndicatorShown(True)
        self.horizontalHeader().setMinimumSectionSize(48)
        self.horizontalHeader().sectionClicked.connect(self._header_clicked)
        self.horizontalHeader().sectionResized.connect(lambda *_: QTimer.singleShot(0, self.adjust_row_heights))
        self.itemClicked.connect(self._copy_item_text)
        self.itemDoubleClicked.connect(self._double_clicked)
        self.itemSelectionChanged.connect(self._selection_changed)
        self.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.customContextMenuRequested.connect(self._open_context_menu)
        self.setHorizontalScrollMode(QAbstractItemView.ScrollMode.ScrollPerPixel)
        self.setVerticalScrollMode(QAbstractItemView.ScrollMode.ScrollPerPixel)
        self._center_delegate = CenteredItemDelegate(self)
        self.setItemDelegate(self._center_delegate)
        self.verticalScrollBar().setSingleStep(8)
        self.horizontalScrollBar().setSingleStep(8)
        self.configure_columns()
        self.setStyleSheet("QTableWidget::item { padding: 4px; }")
        self.setFocusPolicy(Qt.FocusPolicy.StrongFocus)

    def wheelEvent(self, event):
        if event.modifiers() & Qt.KeyboardModifier.ControlModifier:
            super().wheelEvent(event)
            return
        delta = event.pixelDelta().y() if not event.pixelDelta().isNull() else event.angleDelta().y()
        if delta == 0:
            super().wheelEvent(event)
            return
        bar = self.verticalScrollBar()
        bar.setValue(_smooth_wheel_step(delta, bar.value()))
        event.accept()

    def configure_columns(self):
        self.setColumnWidth(0, 40)
        self.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Fixed)
        for idx, (_, _, width, _) in enumerate(self._columns, start=1):
            self.horizontalHeader().setSectionResizeMode(idx, QHeaderView.ResizeMode.Interactive)
            self.setColumnWidth(idx, width)

    def clear_table(self):
        self.setSortingEnabled(False)
        self.blockSignals(True)
        self.clearContents()
        self.setRowCount(0)
        self._rows = []
        self.blockSignals(False)
        self.clear_sort_indicator()
        self.current_cell = (-1, -1)
        self._clear_flash_cells()

    def load_rows(self, rows):
        self.clear_table()
        self._rows = list(rows)
        self.setRowCount(len(self._rows))
        self.blockSignals(True)
        for r_idx, row in enumerate(self._rows):
            serial = QTableWidgetItem(str(r_idx + 1))
            serial.setData(Qt.UserRole, row.get("_db_id"))
            serial.setTextAlignment(Qt.AlignCenter)
            self.setItem(r_idx, 0, serial)

            for c_idx, (cid, _, _, wrap) in enumerate(self._columns, start=1):
                text = str(row.get(cid, "") or "")
                item = QTableWidgetItem(text)
                item.setData(Qt.UserRole, row.get("_db_id"))
                if cid in {"system", "type"}:
                    item.setTextAlignment(Qt.AlignCenter)
                elif cid in {"desc", "solution", "remarks"} or wrap:
                    item.setTextAlignment(Qt.AlignLeft | Qt.AlignVCenter)
                else:
                    item.setTextAlignment(Qt.AlignCenter)
                if cid == "status":
                    fg = {"Open": C["open"], "Clarification": C["clarif"], "Closed": C["closed"]}.get(text, C["text"])
                elif row.get("status") == "Open":
                    fg = C["open"]
                elif row.get("status") == "Clarification":
                    fg = C["clarif"]
                elif row.get("status") == "Closed":
                    fg = C["closed"]
                else:
                    fg = C["text"]
                item.setForeground(QColorFromHex(fg))
                self.setItem(r_idx, c_idx, item)
        self.blockSignals(False)
        self.adjust_row_heights()
        self.clearSelection()
        self.setSortingEnabled(False)
        self.viewport().update()

    def adjust_row_heights(self):
        fm = QFontMetrics(self.font())
        left_align_cols = {"desc", "solution", "remarks"}
        for r_idx, row in enumerate(self._rows):
            row_h = 32
            for c_idx, (cid, _, _, wrap) in enumerate(self._columns, start=1):
                item = self.item(r_idx, c_idx)
                if not item:
                    continue
                text = item.text()
                if not text:
                    continue
                width = max(30, self.columnWidth(c_idx) - 12)
                if cid in left_align_cols and wrap:
                    rect = fm.boundingRect(0, 0, width, 10000, Qt.TextWordWrap, text)
                    h = rect.height() + 14
                else:
                    h = fm.height() + 14
                row_h = max(row_h, h)
            self.setRowHeight(r_idx, row_h)

    def _header_clicked(self, section):
        if section == 0:
            return
        col_id = self._columns[section - 1][0]
        self.sortRequested.emit(col_id)

    def set_sort_indicator(self, col_id=None, reverse=False):
        header = self.horizontalHeader()
        if not col_id:
            header.setSortIndicatorShown(False)
            return
        section = self._col_index_by_id.get(col_id)
        if section is None:
            header.setSortIndicatorShown(False)
            return
        header.setSortIndicatorShown(True)
        header.setSortIndicator(
            section,
            Qt.SortOrder.DescendingOrder if reverse else Qt.SortOrder.AscendingOrder,
        )

    def clear_sort_indicator(self):
        self.set_sort_indicator(None, False)

    def _copy_item_text(self, item):
        if item is not None and item.text():
            self.current_cell = (item.row(), item.column())
            QApplication.clipboard().setText(item.text())
            self.viewport().update()

    def _flash_cell(self, row, column):
        self._flash_cells = {(row, column)}
        self.viewport().update()
        self._flash_timer.start(220)

    def _clear_flash_cells(self):
        if self._flash_cells:
            self._flash_cells.clear()
            self.viewport().update()

    def keyPressEvent(self, event):
        if event.matches(QKeySequence.StandardKey.Copy):
            item = self.currentItem()
            if item is not None and item.text():
                self.current_cell = (item.row(), item.column())
                QApplication.clipboard().setText(item.text())
                self._flash_cell(item.row(), item.column())
                self.viewport().update()
                return
        super().keyPressEvent(event)

    def _double_clicked(self, item):
        if item is None:
            return
        row = item.row()
        db_id = self._db_id_for_row(row)
        if db_id:
            self.editRequested.emit(db_id)

    def _open_context_menu(self, pos):
        item = self.itemAt(pos)
        if item is None:
            return
        row = item.row()
        db_id = self._db_id_for_row(row)
        if db_id is None:
            return

        menu = QMenu(self)
        edit_action = menu.addAction("Edit issue")
        delete_action = menu.addAction("Delete issue")
        chosen = menu.exec(self.viewport().mapToGlobal(pos))
        if chosen == edit_action:
            self.editRequested.emit(db_id)
        elif chosen == delete_action:
            self.deleteRequested.emit(db_id)

    def _selection_changed(self):
        self.selectionChangedCount.emit(len(self.selectedRows()))

    def selectedRows(self):
        rows = []
        for idx in self.selectionModel().selectedRows():
            rows.append(idx.row())
        return rows

    def selected_db_ids(self):
        ids = []
        for row in self.selectedRows():
            db_id = self._db_id_for_row(row)
            if db_id is not None:
                ids.append(db_id)
        return ids

    def _db_id_for_row(self, row):
        item = self.item(row, 0)
        if item is None:
            return None
        return item.data(Qt.UserRole)


def QColorFromHex(hex_color):
    from PySide6.QtGui import QColor

    return QColor(hex_color)


class CenteredItemDelegate(QStyledItemDelegate):
    def paint(self, painter, option, index):
        super().paint(painter, option, index)
        table = self.parent()
        if table is None:
            return
        cell_key = (index.row(), index.column())
        rect = option.rect.adjusted(1, 1, -1, -1)
        painter.save()
        painter.setRenderHint(QPainter.RenderHint.Antialiasing, True)
        if cell_key == getattr(table, "current_cell", (-1, -1)):
            border_color = table._flash_border_color if cell_key in getattr(table, "_flash_cells", set()) else QColor(C["accent"])
            painter.setPen(QPen(border_color, 2))
            painter.drawRoundedRect(rect, 6, 6)
        painter.restore()


def _smooth_wheel_step(delta_y, current_value):
    if delta_y == 0:
        return current_value
    # Smaller-than-default wheel increments feel noticeably smoother.
    step = max(6, int(abs(delta_y) / 4))
    return current_value - step if delta_y > 0 else current_value + step


class SmoothScrollArea(QScrollArea):
    def wheelEvent(self, event):
        if event.modifiers() & Qt.KeyboardModifier.ControlModifier:
            super().wheelEvent(event)
            return
        delta = event.pixelDelta().y() if not event.pixelDelta().isNull() else event.angleDelta().y()
        if delta == 0:
            super().wheelEvent(event)
            return
        bar = self.verticalScrollBar()
        bar.setValue(_smooth_wheel_step(delta, bar.value()))
        event.accept()


class TrendChartWidget(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self._data = None
        self.setMinimumHeight(460)

    def set_data(self, data):
        self._data = data or None
        self.update()

    def _category_order(self, dimension, data):
        if dimension == "issue_type":
            base = ISSUE_TYPES
        else:
            base = FAMILIES
        extras = []
        if data:
            for bucket in data.get("totals", {}).keys():
                if bucket not in base and bucket not in extras:
                    extras.append(bucket)
        return base + extras

    def _palette_for(self, categories):
        colors = {}
        for idx, category in enumerate(categories):
            if category in TREND_ISSUE_COLORS:
                colors[category] = QColor(TREND_ISSUE_COLORS[category])
            else:
                colors[category] = QColor(BULK_IMPORT_COLORS[idx % len(BULK_IMPORT_COLORS)])
        return colors

    def paintEvent(self, event):
        painter = QPainter(self)
        painter.setRenderHint(QPainter.RenderHint.Antialiasing, True)
        painter.fillRect(self.rect(), QColor(C["surface"]))

        data = self._data
        if not data:
            painter.setPen(QColor(C["subtle"]))
            painter.drawText(self.rect(), Qt.AlignCenter, "No trend data available.")
            return

        title = "Issue Trend"
        if data["dimension"] == "product_family":
            title = "Product Family Trend"
        subtitle = "Year overview"
        scope_bits = []
        if data.get("tracker_scope") and data["tracker_scope"] != "All":
            scope_bits.append(data["tracker_scope"])
        if data.get("year_scope") and data["year_scope"] != "All":
            scope_bits.append(data["year_scope"])
        if data.get("month_scope") and data["month_scope"] != "All":
            scope_bits.append(data["month_scope"])
        if scope_bits:
            subtitle = f"{subtitle} | " + " | ".join(scope_bits)

        fm = painter.fontMetrics()
        outer = self.rect().adjusted(18, 18, -18, -18)
        title_font = QFont(F["family"], F["size_lg"])
        title_font.setBold(True)
        painter.setPen(QColor(C["text"]))
        painter.setFont(title_font)
        painter.drawText(QRectF(outer.left(), outer.top(), outer.width(), 26), Qt.AlignLeft | Qt.AlignVCenter, title)
        painter.setFont(QFont(F["family"], F["size_sm"]))
        painter.setPen(QColor(C["subtle"]))
        painter.drawText(QRectF(outer.left(), outer.top() + 24, outer.width(), 22), Qt.AlignLeft | Qt.AlignVCenter, subtitle)

        chart_rect = QRectF(outer.left(), outer.top() + 58, outer.width(), outer.height() - 58)
        painter.setPen(QPen(QColor(C["border"]), 1))
        painter.setBrush(QColor(11, 18, 32, 220))
        painter.drawRoundedRect(chart_rect, 18, 18)

        inner = chart_rect.adjusted(20, 20, -20, -20)
        categories = self._category_order(data["dimension"], data)
        palette = self._palette_for(categories)

        legend_items = [c for c in categories if data.get("totals", {}).get(c, 0)]
        if not legend_items:
            legend_items = categories[:]

        legend_y = inner.top()
        legend_x = inner.left()
        legend_row_h = 20
        legend_gap = 18
        legend_row_gap = 6
        legend_rows = 1
        line_bottom = legend_y + legend_row_h
        for category in legend_items:
            label_w = fm.horizontalAdvance(category)
            item_w = 12 + 8 + label_w
            if legend_x > inner.left() and (legend_x + item_w) > inner.right():
                legend_x = inner.left()
                legend_y += legend_row_h + legend_row_gap
                legend_rows += 1
                line_bottom = legend_y + legend_row_h

            swatch = QRectF(legend_x, legend_y + 4, 12, 12)
            painter.setBrush(palette.get(category, QColor(C["subtle"])))
            painter.setPen(Qt.PenStyle.NoPen)
            painter.drawRoundedRect(swatch, 3, 3)
            painter.setPen(QColor(C["text"]))
            label_rect = QRectF(legend_x + 20, legend_y, item_w - 20, legend_row_h)
            painter.drawText(label_rect, Qt.AlignLeft | Qt.AlignVCenter, category)
            legend_x += item_w + legend_gap

        plot_top = line_bottom + 16
        plot_rect = QRectF(inner.left(), plot_top, inner.width(), inner.bottom() - plot_top)
        if data["mode"] == "year":
            self._paint_year(painter, plot_rect, data, categories, palette)
        else:
            self._paint_month(painter, plot_rect, data, categories, palette)

    def _paint_year(self, painter, rect, data, categories, palette):
        months = [data["months"].get(i, {}) for i in range(1, 13)]
        month_totals = [sum(month.values()) for month in months]
        max_total = max(month_totals) if month_totals else 0
        if max_total <= 0:
            painter.setPen(QColor(C["subtle"]))
            painter.drawText(rect.toRect(), Qt.AlignCenter, "No issues found for the selected filters.")
            return

        left = rect.left() + 36
        top = rect.top() + 10
        right = rect.right() - 10
        bottom = rect.bottom() - 30
        chart = QRectF(left, top, right - left, bottom - top)
        axis_pen = QPen(QColor(C["border"]), 1)
        grid_pen = QPen(QColor("#20314f"), 1)
        painter.setPen(axis_pen)
        painter.setBrush(Qt.BrushStyle.NoBrush)
        painter.drawRect(chart)

        steps = 4
        for step_idx in range(steps + 1):
            y = chart.bottom() - (chart.height() * step_idx / steps)
            value = int(max_total * step_idx / steps)
            painter.setPen(grid_pen)
            painter.drawLine(chart.left(), y, chart.right(), y)
            painter.setPen(QColor(C["subtle"]))
            painter.drawText(QRectF(rect.left(), y - 10, 30, 18), Qt.AlignRight | Qt.AlignVCenter, str(value))

        month_w = chart.width() / 12.0
        active_categories = [c for c in categories if any(month.get(c, 0) for month in months)]
        if not active_categories:
            active_categories = categories[:]
        group_w = month_w * 0.76
        gap = 3
        bar_w = max(8, (group_w - gap * max(0, len(active_categories) - 1)) / max(1, len(active_categories)))
        for month_idx, month_data in enumerate(months, start=1):
            group_left = chart.left() + (month_idx - 1) * month_w + (month_w - group_w) / 2
            total = sum(month_data.values())
            if total <= 0:
                painter.setPen(QColor(C["subtle"]))
                painter.drawText(QRectF(group_left - 6, chart.bottom() + 6, group_w + 12, 18), Qt.AlignCenter, TREND_MONTHS[month_idx - 1])
                continue

            tallest = 0
            for cat_idx, category in enumerate(active_categories):
                count = month_data.get(category, 0)
                if not count:
                    continue
                segment_h = chart.height() * (count / max_total)
                tallest = max(tallest, segment_h)
                x = group_left + cat_idx * (bar_w + gap)
                seg_rect = QRectF(x, chart.bottom() - segment_h, bar_w, segment_h)
                painter.setPen(QPen(QColor("#1d2b44"), 1))
                painter.setBrush(palette.get(category, QColor(C["subtle"])))
                painter.drawRoundedRect(seg_rect, 4, 4)
                if segment_h >= 16:
                    painter.setPen(QColor(C["bg"]))
                    painter.drawText(seg_rect, Qt.AlignCenter, str(count))

            painter.setPen(QColor(C["text"]))
            label_x = group_left - 6
            painter.drawText(QRectF(label_x, chart.bottom() + 6, group_w + 12, 18), Qt.AlignCenter, TREND_MONTHS[month_idx - 1])
            painter.setPen(QColor(C["subtle"]))
            painter.drawText(QRectF(label_x, chart.bottom() - tallest - 22, group_w + 12, 18), Qt.AlignCenter, str(total))

    def _paint_month(self, painter, rect, data, categories, palette):
        buckets = data.get("buckets", {})
        ordered = [c for c in categories if buckets.get(c, 0)]
        if not ordered:
            ordered = categories
        max_total = max([buckets.get(c, 0) for c in ordered] or [0])
        if max_total <= 0:
            painter.setPen(QColor(C["subtle"]))
            painter.drawText(rect.toRect(), Qt.AlignCenter, "No issues found for the selected filters.")
            return

        label_w = 180
        value_w = 60
        bar_left = rect.left() + label_w
        bar_right = rect.right() - value_w
        row_h = max(28, rect.height() / max(1, len(ordered)))
        axis_pen = QPen(QColor(C["border"]), 1)
        painter.setPen(axis_pen)
        painter.setBrush(Qt.BrushStyle.NoBrush)
        painter.drawRect(QRectF(bar_left, rect.top(), bar_right - bar_left, min(rect.height(), row_h * len(ordered))))

        for idx, category in enumerate(ordered):
            y = rect.top() + idx * row_h + 5
            value = buckets.get(category, 0)
            painter.setPen(QColor(C["text"]))
            painter.drawText(QRectF(rect.left(), y, label_w - 14, 18), Qt.AlignRight | Qt.AlignVCenter, category)
            bar_w = (bar_right - bar_left) * (value / max_total)
            bar_rect = QRectF(bar_left, y, bar_w, min(20, row_h - 10))
            painter.setPen(QPen(QColor("#1d2b44"), 1))
            painter.setBrush(palette.get(category, QColor(C["subtle"])))
            painter.drawRoundedRect(bar_rect, 5, 5)
            painter.setPen(QColor(C["subtle"]))
            painter.drawText(QRectF(bar_right + 6, y, value_w - 6, 18), Qt.AlignLeft | Qt.AlignVCenter, str(value))


class TrendShowcaseDialog(QDialog):
    def __init__(self, parent, tracker_scope="All"):
        super().__init__(parent)
        self.setWindowTitle("Trend Showcase")
        self.setModal(True)
        self.resize(1180, 780)
        self.setWindowState(self.windowState() | Qt.WindowState.WindowMaximized)
        self._default_tracker_scope = tracker_scope if tracker_scope in TRACKER_TYPES else "All"
        self._build()
        self.refresh_chart()

    def _build(self):
        self.setStyleSheet(_stylesheet())
        outer = QVBoxLayout(self)
        outer.setContentsMargins(16, 16, 16, 16)
        outer.setSpacing(12)

        header = _card("Trend Showcase")
        header_lay = header.layout()
        header_lay.addWidget(_mk_label("See issue volume by month, grouped by issue type or product family.", color=C["subtle"]))
        outer.addWidget(header)

        control_card = _card("Controls")
        control_lay = QGridLayout()
        control_lay.setContentsMargins(0, 0, 0, 0)
        control_lay.setHorizontalSpacing(10)
        control_lay.setVerticalSpacing(10)
        control_card.layout().addLayout(control_lay)

        self.scope_combo = self._combo(["All"] + TRACKER_TYPES)
        self.scope_combo.setCurrentText(self._default_tracker_scope)
        self.year_combo = self._combo(["All"] + _available_issue_years())
        self.dimension_combo = self._combo(["Issue Type", "Product Family"])
        current_year = str(datetime.date.today().year)
        self.year_combo.setCurrentText(current_year)

        for combo in [self.scope_combo, self.year_combo, self.dimension_combo]:
            combo.currentTextChanged.connect(self.refresh_chart)

        self._wire(control_lay, "Tracker Scope", self.scope_combo, 0)
        self._wire(control_lay, "Year", self.year_combo, 2)
        self._wire(control_lay, "Breakdown", self.dimension_combo, 4)
        self.refresh_btn = _mk_button("Refresh", self.refresh_chart, outline=True, width=100)
        control_lay.addWidget(self.refresh_btn, 0, 6)
        control_lay.setColumnStretch(7, 1)
        outer.addWidget(control_card)

        summary = QHBoxLayout()
        summary.setSpacing(10)
        self.total_card = _metric_card("Total", "0", accent=C["accent"])
        self.peak_card = _metric_card("Peak Month", "-", accent=C["open"])
        self.top_card = _metric_card("Top Issue Type", "-", accent=C["closed"])
        for card in [self.total_card, self.peak_card, self.top_card]:
            summary.addWidget(card)
        outer.addLayout(summary)

        self.chart = TrendChartWidget(self)
        outer.addWidget(self.chart, 1)

    def _combo(self, items):
        combo = FilterDropdown(items, self)
        combo.setMinimumWidth(130)
        return combo

    def _wire(self, row, label, combo, col):
        row.addWidget(_mk_label(label, color=C["subtle"]), 0, col)
        row.addWidget(combo, 0, col + 1)

    def refresh_chart(self, *_):
        dimension = "issue_type" if self.dimension_combo.currentText() == "Issue Type" else "product_family"
        tracker_scope = self.scope_combo.currentText()
        year_scope = self.year_combo.currentText()
        data = fetch_trend_data(tracker_scope, year_scope, "All", dimension)
        self.chart.set_data(data)

        totals = data.get("totals", {})
        total_count = sum(totals.values())
        self.total_card._value_label.setText(str(total_count))

        month_totals = data.get("months", {})
        peak_idx = None
        peak_val = 0
        for idx, month_map in month_totals.items():
            val = sum(month_map.values())
            if val > peak_val:
                peak_val = val
                peak_idx = idx
        self.peak_card._value_label.setText(TREND_MONTHS[(peak_idx or 1) - 1] if peak_idx else "-")

        if totals:
            top_bucket = max(totals, key=totals.get)
            self.top_card._value_label.setText(f"{top_bucket} ({totals[top_bucket]})")
            top_color = TREND_ISSUE_COLORS.get(top_bucket, C["closed"])
            self.top_card._value_label.setStyleSheet(f"font-size: 18pt; color: {top_color};")
        else:
            self.top_card._value_label.setText("-")
            self.top_card._value_label.setStyleSheet(f"font-size: 18pt; color: {C['closed']};")


class TabContentFrame(QWidget):
    def __init__(self, parent, tracker_type, on_double_click, on_sort, on_new_issue, on_delete_issue):
        super().__init__(parent)
        self.tracker_type = tracker_type
        self.on_double_click = on_double_click
        self.on_sort = on_sort
        self.on_new_issue = on_new_issue
        self.on_delete_issue = on_delete_issue
        self._all_rows = []
        self._search_text = ""
        self._sort_col = None
        self._sort_rev = False
        self._status_filter = "All"
        self._build()

    def _build(self):
        lay = QVBoxLayout(self)
        lay.setContentsMargins(0, 14, 0, 0)
        lay.setSpacing(14)

        hero = QFrame(self)
        hero.setObjectName("HeroPanel")
        hero.setStyleSheet(
            f"""
            QFrame#HeroPanel {{
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                                            stop:0 {C['surface_2']},
                                            stop:1 {C['header']});
                border: 1px solid {C['border']};
                border-radius: 18px;
            }}
            """
        )
        hero_row = QHBoxLayout(hero)
        hero_row.setContentsMargins(18, 16, 18, 16)
        hero_row.setSpacing(14)

        title_col = QVBoxLayout()
        title_col.setSpacing(4)
        title_col.addWidget(_mk_label(f"{self.tracker_type} Workspace", bold=True))
        subtitle = _mk_label("Filter, search, create, and resolve issues from one control surface.", color=C["subtle"])
        title_col.addWidget(subtitle)
        hero_row.addLayout(title_col, 2)

        self.sel_lbl = _mk_label("", color=C["subtle"])
        hero_row.addWidget(self.sel_lbl, 0, Qt.AlignVCenter)

        self.search = QLineEdit(self)
        self.search.setPlaceholderText("Search every visible field...")
        self.search.setMinimumWidth(180)
        self.search.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        self.search.textChanged.connect(self.search_apply)
        hero_row.addWidget(self.search, 1, Qt.AlignVCenter)

        self.search_match_lbl = _mk_label("", color=C["subtle"])
        hero_row.addWidget(self.search_match_lbl, 0, Qt.AlignVCenter)

        self.search_clear_btn = _mk_button("Clear", self.search_clear, outline=True, width=70)
        self.search_clear_btn.hide()
        hero_row.addWidget(self.search_clear_btn, 0, Qt.AlignVCenter)

        lay.addWidget(hero)

        metrics = QHBoxLayout()
        metrics.setSpacing(10)
        self.metric_total = _metric_card("Total", "0", accent=C["accent"])
        self.metric_open = _metric_card("Open", "0", accent=C["open"])
        self.metric_clarif = _metric_card("Clarification", "0", accent=C["clarif"])
        self.metric_closed = _metric_card("Closed", "0", accent=C["closed"])
        self.metric_open.clicked.connect(lambda: self.apply_status_filter("Open"))
        self.metric_clarif.clicked.connect(lambda: self.apply_status_filter("Clarification"))
        self.metric_closed.clicked.connect(lambda: self.apply_status_filter("Closed"))
        for card in [self.metric_total, self.metric_open, self.metric_clarif, self.metric_closed]:
            metrics.addWidget(card)
        lay.addLayout(metrics)

        filters_card = _card("Filters")
        filters_body = QWidget(filters_card)
        filters_row = QGridLayout(filters_body)
        filters_row.setContentsMargins(0, 0, 0, 0)
        filters_row.setHorizontalSpacing(10)
        filters_row.setVerticalSpacing(8)
        filters_card.layout().addWidget(filters_body)

        self.fs = self._mk_combo(["All", "Open", "Clarification", "Closed"])
        self.ff = self._mk_combo(["All"] + FAMILIES)
        self.fi = self._mk_combo(["All"] + ISSUE_TYPES)
        self.fm = self._mk_combo(MONTH_NAMES)
        self.fy = self._mk_combo(YEAR_OPTS)

        self._wire_filter(filters_row, "Status", self.fs, 0)
        self._wire_filter(filters_row, "Family", self.ff, 2)
        self._wire_filter(filters_row, "Issue", self.fi, 4)
        self._wire_filter(filters_row, "Month", self.fm, 6)
        self._wire_filter(filters_row, "Year", self.fy, 8)

        reset_btn = _mk_button("Reset Filters", self.reset_filters, outline=True, width=132)
        filters_row.addWidget(reset_btn, 0, 10, 1, 1)
        for col in (1, 3, 5, 7, 9, 11):
            filters_row.setColumnStretch(col, 1)
        lay.addWidget(filters_card)

        table_card = _card()
        table_lay = QVBoxLayout()
        table_lay.setContentsMargins(10, 10, 10, 10)
        self.table = IssueTableWidget(self)
        self.table.sortRequested.connect(self._sort_requested)
        self.table.editRequested.connect(self.on_double_click)
        self.table.deleteRequested.connect(self.on_delete_issue)
        self.table.selectionChangedCount.connect(self.update_selection_label)
        table_lay.addWidget(self.table)
        table_card.setLayout(table_lay)
        lay.addWidget(table_card, 1)

    def _mk_combo(self, items):
        field = FilterDropdown(items, self)
        field.currentTextChanged.connect(self._refresh)
        field.setMinimumWidth(96)
        field.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        return field

    def _wire_filter(self, row, label, combo, col):
        row.addWidget(_mk_label(label, color=C["subtle"]), 0, col)
        row.addWidget(combo, 0, col + 1)

    def _refresh(self, *_):
        for combo in [self.fs, self.ff, self.fi, self.fm, self.fy]:
            combo.blockSignals(True)
        self.search.blockSignals(True)
        self.search.setText("")
        self.search.blockSignals(False)
        self.search_match_lbl.setText("")
        self.search_clear_btn.hide()
        for combo in [self.fs, self.ff, self.fi, self.fm, self.fy]:
            combo.blockSignals(False)

        self._sort_col = None
        self._sort_rev = False
        self._status_filter = self.fs.currentText()
        rows = fetch_issues(self.fs.currentText(), self.ff.currentText(), self.fi.currentText(), self.fm.currentText(), self.fy.currentText(), self.tracker_type)
        self._all_rows = []
        for row in rows:
            self._all_rows.append(
                {
                    "_db_id": row["id"],
                    "date": _fmt_display_date(row["date_reported"]),
                    "system": row["system_number"] or "",
                    "family": row["product_family"] or "",
                    "type": row["issue_type"] or "",
                    "sps": row["sps_number"] or "",
                    "ncr": row["ncr_number"] or "",
                    "status": row["status"] or "",
                    "desc": row["issue_desc"] or "",
                    "solution": row["solution"] or "",
                    "sol_date": _fmt_display_date(row["solution_date"]),
                    "crf": row["crf"] or "",
                    "esw": row["esw"] or "",
                    "scv": row["scv"] or "",
                    "remarks": row["remarks"] or "",
                }
            )
        self._apply_visible_rows()
        self.table.clear_sort_indicator()
        total, open_, closed, clarif = get_counts(self.tracker_type)
        self.metric_total._value_label.setText(str(total))
        self.metric_open._value_label.setText(str(open_))
        self.metric_clarif._value_label.setText(str(clarif))
        self.metric_closed._value_label.setText(str(closed))
        self._sync_metric_cards()
        self.sel_lbl.setText("")

    def _apply_visible_rows(self):
        query = self.search.text().strip().lower()
        if query:
            tokens = query.split()
            visible = []
            for row in self._all_rows:
                haystack = " ".join(str(v) for v in row.values()).lower()
                if all(token in haystack for token in tokens):
                    visible.append(row)
        else:
            visible = list(self._all_rows)
        self.table.load_rows(visible)
        if self._sort_col:
            self.table.set_sort_indicator(self._sort_col, self._sort_rev)
        else:
            self.table.clear_sort_indicator()
        if query:
            n = len(visible)
            self.search_match_lbl.setText(f"{n} match{'es' if n != 1 else ''}" if n else "No matches")
            self.search_match_lbl.setStyleSheet(f"color: {C['subtle'] if n else C['open']};")
            self.search_clear_btn.show()
        else:
            self.search_match_lbl.setText("")
            self.search_clear_btn.hide()

    def _sort_requested(self, col_id):
        self._sort_rev = not self._sort_rev if self._sort_col == col_id else False
        self._sort_col = col_id
        self._all_rows.sort(key=lambda r: str(r.get(col_id, "")), reverse=self._sort_rev)
        self._apply_visible_rows()
        self.table.set_sort_indicator(col_id, self._sort_rev)
        self.on_sort(self.tracker_type, col_id, self._sort_rev)

    def reset_filters(self):
        for combo in [self.fs, self.ff, self.fi, self.fm, self.fy]:
            with QSignalBlocker(combo):
                combo.setCurrentText("All")
        self._status_filter = "All"
        self._sync_metric_cards()
        self._refresh()
        self.table.clear_sort_indicator()

    def apply_status_filter(self, status):
        next_status = "All" if self._status_filter == status else status
        self._status_filter = next_status
        with QSignalBlocker(self.fs):
            self.fs.setCurrentText(next_status)
        self._sync_metric_cards()
        self._refresh()

    def _sync_metric_cards(self):
        states = {
            "Open": self.metric_open,
            "Clarification": self.metric_clarif,
            "Closed": self.metric_closed,
        }
        self.metric_total.setProperty("selected", False)
        self.metric_total.style().unpolish(self.metric_total)
        self.metric_total.style().polish(self.metric_total)
        self.metric_total.update()
        for status, card in states.items():
            selected = (self._status_filter == status)
            card.setProperty("selected", selected)
            card.style().unpolish(card)
            card.style().polish(card)
            card.update()

    def search_apply(self, *_):
        self._apply_visible_rows()

    def search_clear(self):
        self.search.clear()
        self.search.setFocus()
        self._apply_visible_rows()

    def get_selected_db_ids(self):
        return self.table.selected_db_ids()

    def update_selection_label(self, n):
        self.sel_lbl.setText(f"{n} row{'s' if n != 1 else ''} selected" if n else "")


class ExportScopeDialog(QDialog):
    def __init__(self, parent):
        super().__init__(parent)
        self.result = None
        self.setWindowTitle("Select Export Scope")
        self.resize(450, 320)
        self.setMinimumSize(400, 280)
        self.setModal(True)
        self._build()

    def _build(self):
        self.setStyleSheet(_stylesheet())
        outer = QVBoxLayout(self)
        outer.setContentsMargins(16, 14, 16, 14)
        outer.setSpacing(8)
        outer.addWidget(_mk_label("Select Export Scope", bold=True))
        outer.addWidget(_mk_label("Choose which issue types to include in the export:", color=C["subtle"]))
        outer.addSpacing(2)

        self.scope = "Both"
        self.rb_eng = QRadioButton("Engineering Issues Only")
        self.rb_mat = QRadioButton("Material Issues Only")
        self.rb_both = QRadioButton("Both Engineering and Material Issues")
        self.rb_both.setChecked(True)
        for rb in (self.rb_eng, self.rb_mat, self.rb_both):
            rb.setCursor(Qt.PointingHandCursor)
            rb.setStyleSheet(
                f"""
                QRadioButton {{
                    background: {C['surface']};
                    color: {C['text']};
                    border: 1px solid {C['border']};
                    border-radius: 12px;
                    padding: 12px 14px;
                    margin: 0px;
                    spacing: 10px;
                    font-weight: 600;
                }}
                QRadioButton:hover {{
                    border: 1px solid {C['accent']};
                    background: {C['sel']};
                }}
                QRadioButton:checked {{
                    border: 1px solid {C['accent']};
                    background: {C['sel']};
                    color: white;
                }}
                QRadioButton::indicator {{
                    width: 14px;
                    height: 14px;
                }}
                QRadioButton::indicator:unchecked {{
                    border: 2px solid {C['subtle']};
                    border-radius: 7px;
                    background: transparent;
                }}
                QRadioButton::indicator:checked {{
                    border: 2px solid {C['accent']};
                    border-radius: 7px;
                    background: {C['accent']};
                }}
                """
            )
        outer.addWidget(self.rb_eng)
        outer.addSpacing(4)
        outer.addWidget(self.rb_mat)
        outer.addSpacing(4)
        outer.addWidget(self.rb_both)
        outer.addStretch(1)

        row = QHBoxLayout()
        row.setContentsMargins(0, 0, 0, 0)
        row.setSpacing(10)
        row.addStretch(1)
        row.addWidget(_mk_button("Cancel", self.reject, outline=True, width=90))
        row.addWidget(_mk_button("Export", self._export, width=120))
        outer.addLayout(row)

    def _export(self):
        if self.rb_eng.isChecked():
            self.result = "Engineering"
        elif self.rb_mat.isChecked():
            self.result = "Material"
        else:
            self.result = "Both"
        self.accept()


class AppWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self._tab_frames = {}
        self._bulk_import_path = None
        self._bulk_import_waiting = False
        self.setWindowTitle(WIN_TITLE)
        self.resize(*WIN_SIZE)
        self.setMinimumSize(860, 560)
        self.setStyleSheet(_stylesheet())
        self._build()
        self.refresh_all_tabs()

    def _build(self):
        central = QWidget(self)
        central.setObjectName("AppRoot")
        self.setCentralWidget(central)
        root = QHBoxLayout(central)
        root.setContentsMargins(16, 16, 16, 16)
        root.setSpacing(14)

        sidebar_scroll = SmoothScrollArea(self)
        sidebar_scroll.setWidgetResizable(True)
        sidebar_scroll.setFrameShape(QFrame.NoFrame)
        sidebar_scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        sidebar_scroll.setSizePolicy(QSizePolicy.Policy.Preferred, QSizePolicy.Policy.Expanding)
        sidebar_scroll.setMinimumWidth(240)

        sidebar = _card()
        sidebar.setMinimumWidth(240)
        side = QVBoxLayout(sidebar)
        side.setContentsMargins(18, 18, 18, 18)
        side.setSpacing(14)

        brand = QFrame()
        brand.setStyleSheet(
            f"""
            QFrame {{
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                                            stop:0 {C['surface_2']},
                                            stop:1 {C['panel']});
                border: none;
                border-radius: 16px;
            }}
            """
        )
        brand_lay = QVBoxLayout(brand)
        brand_lay.setContentsMargins(16, 16, 16, 16)
        brand_lay.setSpacing(8)
        if os.path.exists(LOGO_PATH):
            try:
                pix = QPixmap(LOGO_PATH)
                if not pix.isNull():
                    pix = pix.scaled(72, 52, Qt.KeepAspectRatio, Qt.SmoothTransformation)
                    logo = QLabel()
                    logo.setPixmap(pix)
                    brand_lay.addWidget(logo)
            except Exception:
                pass
        brand_title = _mk_label("Production Issue", bold=True)
        brand_sub = _mk_label("Tracker Studio", bold=True, color=C["accent"])
        brand_desc = _mk_label("A faster way to log, triage, and export issues.", color=C["subtle"])
        brand_desc.setWordWrap(True)
        brand_desc.setMinimumHeight(36)
        brand_lay.addWidget(brand_title)
        brand_lay.addWidget(brand_sub)
        brand_lay.addWidget(brand_desc)
        side.addWidget(brand)

        action_card = _card("Quick Actions")
        action_lay = action_card.layout()
        self.quick_new_btn = _mk_button("New Issue", self._quick_new_issue)
        self.quick_trend_btn = _mk_button("Trend Showcase", self.show_trend_showcase, outline=True)
        self.quick_bulk_import_btn = _mk_button("Bulk Import", self.import_bulk_issues, outline=True)
        self.quick_refresh_btn = _mk_button("Refresh All", self.refresh_all_tabs, outline=True)
        self.quick_export_btn = _mk_button("Export Excel", self.export_data, outline=True)
        action_lay.addWidget(self.quick_new_btn)
        action_lay.addWidget(self.quick_trend_btn)
        action_lay.addWidget(self.quick_bulk_import_btn)
        action_lay.addWidget(self.quick_refresh_btn)
        action_lay.addWidget(self.quick_export_btn)
        side.addWidget(action_card)

        summary_card = _card("Current Snapshot")
        summary_lay = summary_card.layout()
        self.active_tab_lbl = _mk_label("Engineering Issues", bold=True)
        summary_lay.addWidget(self.active_tab_lbl)
        self.side_total_lbl = _mk_label("Total: 0")
        self.side_open_lbl = _mk_label("Open: 0")
        self.side_clarif_lbl = _mk_label("Clarification: 0")
        self.side_closed_lbl = _mk_label("Closed: 0")
        for lab in [self.side_total_lbl, self.side_open_lbl, self.side_clarif_lbl, self.side_closed_lbl]:
            summary_lay.addWidget(lab)
        side.addWidget(summary_card)

        note_card = _card("Tip")
        note_lay = note_card.layout()
        tip_text = _mk_label("Use the search bar to filter the visible tab instantly. Sorting happens by clicking column headers.", color=C["subtle"])
        tip_text.setWordWrap(True)
        tip_text.setMinimumHeight(42)
        note_lay.addWidget(tip_text)
        side.addWidget(note_card)
        side.addStretch(1)
        side.addWidget(_mk_label("AMAT Production Issue Tracker", color=C["subtle"]))
        side.addWidget(_mk_label("Made by Sankar | v4.1", color=C["subtle"]))
        sidebar_scroll.setWidget(sidebar)

        main = QVBoxLayout()
        main.setSpacing(16)

        header = QFrame(self)
        header.setObjectName("HeaderBar")
        header.setMinimumHeight(88)
        header.setStyleSheet(
            f"""
            QFrame#HeaderBar {{
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                                            stop:0 {C['header']},
                                            stop:0.45 {C['surface_2']},
                                            stop:1 {C['panel']});
                border: none;
                border-radius: 18px;
            }}
            """
        )
        h = QHBoxLayout(header)
        h.setContentsMargins(18, 14, 18, 14)
        h.setSpacing(12)

        headline = QVBoxLayout()
        headline.setSpacing(4)
        headline.addWidget(_mk_label(WIN_TITLE, bold=True, object_name="TitleLabel"))
        headline.addWidget(_mk_label("Two focused workspaces for engineering and material issues.", color=C["subtle"]))
        h.addLayout(headline, 2)

        h.addStretch(1)
        self.header_scope_lbl = _mk_label("Active: Engineering", bold=True, color=C["accent"])
        h.addWidget(self.header_scope_lbl)
        main.addWidget(header)

        self.tabs = QTabWidget(self)
        self.tabs.setDocumentMode(True)
        self.tabs.tabBar().setDrawBase(False)
        self.tabs.currentChanged.connect(self._on_tab_changed)
        main.addWidget(self.tabs, 1)

        for tab_name, tracker_type in [("Engineering Workspace", "Engineering"), ("Material Workspace", "Material")]:
            page = QWidget()
            page_lay = QVBoxLayout(page)
            page_lay.setContentsMargins(0, 8, 0, 0)
            frame = TabContentFrame(
                page,
                tracker_type,
                self.open_issue_by_id,
                self.on_table_sort,
                self.new_issue,
                self.delete_issue_by_id,
            )
            page_lay.addWidget(frame, 1)
            self.tabs.addTab(page, tab_name)
            self._tab_frames[tracker_type] = frame

        root.addWidget(sidebar_scroll)
        main_wrapper = QWidget(self)
        main_wrapper.setLayout(main)
        main_wrapper.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        root.addWidget(main_wrapper, 1)

    def current_tracker_type(self):
        return "Engineering" if self.tabs.currentIndex() == 0 else "Material"

    def refresh_all_tabs(self):
        for frame in self._tab_frames.values():
            frame._refresh()
        self._sync_sidebar()

    def _sync_sidebar(self):
        tracker = self.current_tracker_type()
        self.active_tab_lbl.setText(f"{tracker} Workspace")
        self.header_scope_lbl.setText(f"Active: {tracker}")
        total, open_, closed, clarif = get_counts(tracker)
        self.side_total_lbl.setText(f"Total: {total}")
        self.side_open_lbl.setText(f"Open: {open_}")
        self.side_clarif_lbl.setText(f"Clarification: {clarif}")
        self.side_closed_lbl.setText(f"Closed: {closed}")

    def _on_tab_changed(self, _):
        self._sync_sidebar()

    def new_issue(self, tracker_type):
        dlg = NewIssueDialog(self, tracker_type)
        if dlg.exec() == QDialog.Accepted:
            self._tab_frames[tracker_type]._refresh()
        self._sync_sidebar()

    def open_issue_by_id(self, db_id):
        issue = fetch_by_id(db_id)
        if not issue:
            return
        tracker_type = issue.get("tracker_type", "Engineering")
        dlg = ManageDialog(self, db_id)
        dlg.exec()
        if tracker_type in self._tab_frames:
            self._tab_frames[tracker_type]._refresh()
        self._sync_sidebar()

    def delete_issue_by_id(self, db_id):
        issue = fetch_by_id(db_id)
        if not issue:
            return
        tracker_type = issue.get("tracker_type", "Engineering")
        system_title = (issue.get("system_number") or "").splitlines()[0]
        prompt = f"Delete this issue?\n\n{system_title}" if system_title else "Delete this issue?"
        if QMessageBox.question(
            self,
            "Delete Issue",
            prompt,
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No,
        ) == QMessageBox.StandardButton.Yes:
            delete_by_ids([db_id])
            if tracker_type in self._tab_frames:
                self._tab_frames[tracker_type]._refresh()
            self._sync_sidebar()

    def mass_delete(self, tracker_type):
        frame = self._tab_frames[tracker_type]
        ids = frame.get_selected_db_ids()
        if not ids:
            QMessageBox.information(self, "No Selection", "Select rows to delete first.")
            return
        if QMessageBox.question(
            self,
            "Confirm Delete",
            f"Permanently delete {len(ids)} {tracker_type} issue(s)?\nThis cannot be undone.",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No,
        ) == QMessageBox.StandardButton.Yes:
            delete_by_ids(ids)
            frame._refresh()
        self._sync_sidebar()

    def on_table_sort(self, tracker_type, col_id, reverse):
        return

    def export_data(self):
        export_dialog = ExportScopeDialog(self)
        if export_dialog.exec() != QDialog.Accepted or not export_dialog.result:
            return

        export_scope = export_dialog.result
        active_tracker = self.current_tracker_type()
        tab_frame = self._tab_frames[active_tracker]

        engineering_rows = []
        material_rows = []
        if export_scope in ["Engineering", "Both"]:
            engineering_rows = fetch_issues(
                status_f=tab_frame.fs.currentText(),
                family_f=tab_frame.ff.currentText(),
                itype_f=tab_frame.fi.currentText(),
                month_f=tab_frame.fm.currentText(),
                year_f=tab_frame.fy.currentText(),
                tracker_type="Engineering",
            )
        if export_scope in ["Material", "Both"]:
            material_rows = fetch_issues(
                status_f=tab_frame.fs.currentText(),
                family_f=tab_frame.ff.currentText(),
                itype_f=tab_frame.fi.currentText(),
                month_f=tab_frame.fm.currentText(),
                year_f=tab_frame.fy.currentText(),
                tracker_type="Material",
            )

        customer_mode = False
        if tab_frame.fs.currentText() == "Open":
            answer = QMessageBox.question(
                self,
                "Export Type",
                "You are exporting Open issues.\n\nExport as Customer Report?\n(Yes = customer view, no internal fields)\n(No = full internal log)",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No | QMessageBox.StandardButton.Cancel,
                QMessageBox.StandardButton.No,
            )
            if answer == QMessageBox.StandardButton.Cancel:
                return
            customer_mode = answer == QMessageBox.StandardButton.Yes

        suffix = "_Customer" if customer_mode else ""
        scope_suffix = f"_{export_scope}" if export_scope != "Both" else ""
        initial = f"AMAT_Issues{scope_suffix}{suffix}_{datetime.date.today()}.xlsx"
        path, _ = QFileDialog.getSaveFileName(self, "Save Excel", initial, "Excel Workbook (*.xlsx)")
        if not path:
            return

        try:
            export_excel(engineering_rows, material_rows, path, customer_mode=customer_mode, export_scope=export_scope)
            total_rows = len(engineering_rows) + len(material_rows)
            if QMessageBox.question(
                self,
                "Exported",
                f"Saved {total_rows} rows.\n\nOpen file now?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.No,
            ) == QMessageBox.StandardButton.Yes:
                try:
                    os.startfile(path)
                except Exception:
                    pass
        except Exception as e:
            QMessageBox.critical(self, "Export Error", str(e))

    def _quick_new_issue(self):
        self.new_issue(self.current_tracker_type())

    def show_trend_showcase(self):
        dlg = TrendShowcaseDialog(self, tracker_scope=self.current_tracker_type())
        dlg.exec()

    def create_bulk_template_file(self):
        path = _bulk_import_template_path()
        try:
            if os.path.exists(path):
                try:
                    os.remove(path)
                except Exception:
                    pass
            create_bulk_import_template(path)
            try:
                os.startfile(path)
            except Exception:
                pass
            QMessageBox.information(
                self,
                "Bulk Import Template",
                "A fresh bulk import template has been created and opened.\n\n"
                "Fill the sheet, save it, then close Excel to continue the import prompt.",
            )
        except Exception as e:
            QMessageBox.critical(self, "Template Error", str(e))

    def import_bulk_issues(self):
        if self._bulk_import_waiting:
            return
        self._bulk_import_path = _bulk_import_template_path()
        try:
            if os.path.exists(self._bulk_import_path):
                try:
                    os.remove(self._bulk_import_path)
                except Exception:
                    pass
            create_bulk_import_template(self._bulk_import_path)
            try:
                os.startfile(self._bulk_import_path)
            except Exception as e:
                QMessageBox.critical(self, "Bulk Import Error", f"Could not open the template file.\n\n{e}")
                self._bulk_import_path = None
                return
            self._bulk_import_waiting = True
            QTimer.singleShot(1500, self._poll_bulk_import_template)
        except Exception as e:
            QMessageBox.critical(self, "Bulk Import Error", str(e))
            self._bulk_import_path = None
            self._bulk_import_waiting = False

    def _poll_bulk_import_template(self):
        path = self._bulk_import_path
        if not path:
            self._bulk_import_waiting = False
            return
        if not os.path.exists(path):
            self._bulk_import_waiting = False
            self._bulk_import_path = None
            return
        try:
            with open(path, "r+b"):
                pass
        except Exception:
            QTimer.singleShot(1500, self._poll_bulk_import_template)
            return

        self._bulk_import_waiting = False
        reply = QMessageBox.question(
            self,
            "Import Bulk Issues",
            "The Excel file is closed.\n\nDo you want to import the rows now?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.Yes,
        )
        if reply != QMessageBox.StandardButton.Yes:
            self._bulk_import_path = None
            return

        try:
            imported, skipped = import_bulk_issues(path, default_tracker_type=self.current_tracker_type())
            self.refresh_all_tabs()
            message = f"Imported {imported} issue{'s' if imported != 1 else ''}."
            if skipped:
                preview = "\n".join(skipped[:8])
                if len(skipped) > 8:
                    preview += f"\n... and {len(skipped) - 8} more skipped row(s)."
                message += f"\n\nSkipped {len(skipped)} row(s):\n{preview}"
                QMessageBox.warning(self, "Bulk Import Completed", message)
            else:
                QMessageBox.information(self, "Bulk Import Completed", message)
        except Exception as e:
            QMessageBox.critical(self, "Bulk Import Error", str(e))
        finally:
            self._bulk_import_path = None


def main():
    init_db()
    app = QApplication(sys.argv)
    app.setApplicationName(WIN_TITLE)
    app.setStyle("Fusion")
    app.setFont(QFont(F["family"], F["size_md"]))
    app.setStyleSheet(_stylesheet())
    win = AppWindow()
    win.showMaximized()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
